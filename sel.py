print("Hellossssssssss")
import sys
import time
import os
import shutil
import pandas as pd
from dotenv import load_dotenv
import mysql.connector
from mysql.connector import errorcode
from sqlalchemy import create_engine
from sqlalchemy.exc import SQLAlchemyError
import numpy as np
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
from openpyxl.styles import Font, Color, PatternFill, Border
from openpyxl import Workbook
from openpyxl.comments import Comment


load_dotenv()

ip_addr=os.getenv("IP_ADDRESS")
db_user = os.getenv("DB_USER")
db_password = os.getenv("DB_PASSWORD")
db_database = os.getenv("DB_DATABASE")
dbname = os.getenv("DBNAME")

cnx1=mysql.connector.connect(user=f'{db_user}',password=f'{db_password}',
                   host=f'{ip_addr}',
                   database=dbname)
cnx2=mysql.connector.connect(user=f'{db_user}',password=f'{db_password}',
                   host=f'{ip_addr}',
                   database=f'{os.getenv("DB_DATABASE1")}')


# Construct the URL
database_url = f'mysql+mysqlconnector://{db_user}:{db_password}@{ip_addr}/{db_database}'
# Create an SQLAlchemy Engine with the URL
engine = create_engine(database_url)



# Define a function to establish the MySQL connection
def create_mysql_connection():
    config = {
        "user": db_user,
        "password": db_password,
        "host": ip_addr,  # Or the MySQL server host
        "database": db_database,
    }
    cnx = mysql.connector.connect(**config)
    return cnx

cnx = create_mysql_connection();




def func5(source_table, destination_table, uploaddate, cnx, size=100):
    engine = create_engine(f'mysql+mysqlconnector://{db_user}:{db_password}@{ip_addr}/{db_database}?connect_timeout=60')
    cursor = cnx.cursor()

    try:
        # Connect to the database using the SQLAlchemy engine
        with engine.connect() as connection:
            # Begin a transaction
            trans = connection.begin()

            try:
                # Delete records from the destination table with matching uploaddate
                delete_query = f"DELETE FROM {destination_table} WHERE uploaddate = %s"
                cursor.execute(delete_query, (uploaddate,))
                cnx.commit()
                print('Delete completed')

                # Execute the query to count the number of rows
                count_query = f"SELECT COUNT(*) FROM {source_table}"
                cursor.execute(count_query)

                # Fetch the result
                total_rows = cursor.fetchone()[0]

                # Print the result
                print(f"Total rows : {total_rows}")

                # Read data and insert
                for start_index in range(0, total_rows, size):
                    end_index = start_index + size  # Fix the end_index 

                    # Read a size of data from the source table into a DataFrame
                    query = f"SELECT * FROM {source_table} LIMIT {start_index}, {size}"
                    df = pd.read_sql(query, con=connection)

                    # Drop the index column from the DataFrame if it exists
                    df = df.drop(columns=df.columns[df.columns.str.contains('Unnamed')])
                    print(df.shape)

                    # Add the uploaddate column to the DataFrame
                    df['uploaddate'] = uploaddate


                    # Copy data to the destination table using DataFrame's to_sql method
                    df.to_sql(destination_table, con=connection, if_exists='append', index=False)

                    print(f"Data copied from '{source_table}' to '{destination_table}' for rows {start_index + 1} to {end_index}.")
                    
            # Commit the transaction for each chunk
                trans.commit()

                print(f"Data copied from '{source_table}' to '{destination_table}' successfully for uploaddate '{uploaddate}'.")
                print(f"Source table '{source_table}' dropped.")

            except Exception as e:
                # Rollback the transaction if an error occurs
                trans.rollback()
                print(f"Error: {e}")

    except Exception as conn_err:
        print(f"Connection Error: {conn_err}")




        
def DownloadColumnsFromTo(driver, from_col, to_col, column_line, filename):
    editcolumns_button = driver.find_element(By.PARTIAL_LINK_TEXT, "EDIT COLUMNS").click()
    time.sleep(5)
    reset_defaults = driver.find_element(By.ID, "manage-reset").click()
    for j in range(8):
        path_j = '//*[@id="manage-menu"]/li[1]/button/i'
        driver.find_element(By.XPATH, path_j).click()
   
    search_box = driver.find_element(By.ID, "manage-search")
    search_box.send_keys(' ')
   
    for i in range(from_col, to_col+1):
       
        path_i = f'/html/body/main/div[2]/form/div[3]/div[{column_line}]/label[{i}]/input'
        driver.find_element(By.XPATH, path_i).click()
   
    try:
        element_H1 = driver.find_element(By.XPATH, '/html/body/main/div[2]/form/h1')
        driver.execute_script("arguments[0].scrollIntoView(true);", element_H1)
        driver.find_element(By.XPATH, '/html/body/main/div[2]/form/div[1]/div/button').click()
        driver.find_element(By.XPATH, '/html/body/main/div[2]/div[4]/div[2]/form/button').click()
        time.sleep(5)
       
       
    except Exception as e:
        print(e)



def func2(from_path, to_path):
    if not os.path.exists(to_path):
        os.mkdir(to_path)

    shutil.copy2(os.path.join(from_path, "marketcap.csv"), os.path.join(to_path, f"marketcap_0_{dateObj}.csv"))
    os.remove(os.path.join(from_path, "marketcap.csv"))

    shutil.copy2(os.path.join(from_path, "marketcap (1).csv"), os.path.join(to_path, f"marketcap_1_{dateObj}.csv"))
    os.remove(os.path.join(from_path, "marketcap (1).csv"))

    shutil.copy2(os.path.join(from_path, "marketcap (2).csv"), os.path.join(to_path, f"marketcap_2_{dateObj}.csv"))
    os.remove(os.path.join(from_path, "marketcap (2).csv"))

    shutil.copy2(os.path.join(from_path, "marketcap (3).csv"), os.path.join(to_path, f"marketcap_3_{dateObj}.csv"))
    os.remove(os.path.join(from_path, "marketcap (3).csv"))

    shutil.copy2(os.path.join(from_path, "marketcap (4).csv"), os.path.join(to_path, f"marketcap_4_{dateObj}.csv"))
    os.remove(os.path.join(from_path, "marketcap (4).csv"))

    shutil.copy2(os.path.join(from_path, "marketcap (5).csv"), os.path.join(to_path, f"marketcap_5_{dateObj}.csv"))
    os.remove(os.path.join(from_path, "marketcap (5).csv"))

    shutil.copy2(os.path.join(from_path, "marketcap (6).csv"), os.path.join(to_path, f"marketcap_6_{dateObj}.csv"))
    os.remove(os.path.join(from_path, "marketcap (6).csv"))

    shutil.copy2(os.path.join(from_path, "marketcap (7).csv"), os.path.join(to_path, f"marketcap_7_{dateObj}.csv"))
    os.remove(os.path.join(from_path, "marketcap (7).csv"))

    shutil.copy2(os.path.join(from_path, "marketcap (8).csv"), os.path.join(to_path, f"marketcap_8_{dateObj}.csv"))
    os.remove(os.path.join(from_path, "marketcap (8).csv"))


def func1(to_path):
    chrome_driver_path = os.getenv("CHROME_DRIVER_PATH")
    service = Service(chrome_driver_path)
    service.start()

    options = Options()

    driver = webdriver.Remote(service.service_url, options=options)
    #driver = webdriver.Chrome(chrome_driver_path, options=options)


    driver.get("https://www.screener.in/login/?")
    driver.maximize_window()

    id_username = driver.find_element(By.XPATH, '//*[@id="id_username"]').send_keys(os.getenv("RECIPIENT_EMAIL"))
    id_password = driver.find_element(By.XPATH, '//*[@id="id_password"]').send_keys(os.getenv("ID_PASSWORD"))
    
    login_button = driver.find_element(By.XPATH, '/html/body/main/div[2]/div[2]/form/button').click()
    screens_button = driver.find_element(By.PARTIAL_LINK_TEXT, "SCREENS").click()
    marketcap_button = driver.find_element(By.PARTIAL_LINK_TEXT, "marketcap").click()


    if not os.path.exists(to_path):
        os.mkdir(to_path)

    DownloadColumnsFromTo(driver, 1, 50, 1, f"marketcap_0_{dateObj}")
    DownloadColumnsFromTo(driver, 51, 100, 1, f"marketcap_1_{dateObj}")
    DownloadColumnsFromTo(driver, 101, 150, 1, f"marketcap_2_{dateObj}")
    DownloadColumnsFromTo(driver, 151, 197, 1, f"marketcap_3_{dateObj}")
    DownloadColumnsFromTo(driver, 1, 50, 2, f"marketcap_4_{dateObj}")
    DownloadColumnsFromTo(driver, 50, 51, 2, f"marketcap_5_{dateObj}")
    DownloadColumnsFromTo(driver, 1, 50, 3, f"marketcap_6_{dateObj}")
    DownloadColumnsFromTo(driver, 51, 100, 3, f"marketcap_7_{dateObj}")
    DownloadColumnsFromTo(driver, 101, 110, 3, f"marketcap_8_{dateObj}")
    time.sleep(5)  # Let the user actually see something!




#while True:
#    pass  # Keeps the script running until manually terminated


def func3(directory, output_file):
    files = [file for file in os.listdir(directory) if file.endswith('.csv')]
    dfs = []

    for file in files:
        file_path = os.path.join(directory, file)
        df = pd.read_csv(file_path)
        dfs.append(df)

    merged_df = pd.concat(dfs, axis=1)
    merged_df.to_csv(os.path.join(directory, output_file), index=False)
    print(f"CSV files merged successfully to {output_file}")


def create_stockinfo_table2(csv_file, table_name,cnx):
    # Read the CSV file
    df = pd.read_csv(csv_file)
    print('read csv')
    # Convert column names to lowercase and replace spaces with underscore
    df.columns = df.columns.str.lower().str.replace(' ', '_')
    df.columns = df.columns.str.replace('.', '_')

    # Remove any duplicate columns
    df = df.loc[:, ~df.columns.duplicated()]

    # Get the maximum length of each column
    column_lengths = df.astype(str).apply(lambda x: x.str.len()).max().tolist()

    # Limit column names to 64 characters
    df.columns = df.columns.str[:64]

    # Replace NaN values with None (null values in Python)
    df = df.where(pd.notna(df), None)

    # Convert numeric columns to float, replacing any non-convertible values with None
    numeric_cols = df.select_dtypes(include=[np.number]).columns
    df[numeric_cols] = df[numeric_cols].apply(pd.to_numeric, errors='coerce')

    # Add the 'uploaddate' column with the format 'yyyymmdd'
    dateObj = pd.to_datetime('today').strftime("%Y%m%d")
    df['uploaddate'] = dateObj  # Add the uploaddate column with the same date for all rows

    # Create the stockinfo table in MySQL
    cursor = cnx.cursor()
    cursor.execute(f"DROP TABLE IF EXISTS {table_name}")

    create_table_query = f"CREATE TABLE {table_name} ("
    for col, length in zip(df.columns, column_lengths):
        col_data_type = 'float' if df[col].dtype == float else f"varchar({length})"
        create_table_query += f"{col} {col_data_type}, "
    create_table_query += "uploaddate varchar(8))"

    cursor.execute(create_table_query)

    # Insert data into the stockinfo table
    for row in df.itertuples(index=False, name=None):
        row_without_nan = [None if pd.isna(value) else value for value in row]
        insert_query = f"INSERT INTO {table_name} VALUES ({', '.join(['%s'] * len(row_without_nan))})"
        cursor.execute(insert_query, row_without_nan)

    # Commit changes and close the connection
    cnx.commit()
    cursor.close()

    print(f"Stockinfo table '{table_name}' created successfully.")

def create_stockinfo_table1(csv_file, table_name,cnx):
    # Read the CSV file
    df = pd.read_csv(csv_file)
    print('read csv')
    # Convert column names to lowercase and replace spaces with underscore
    df.columns = df.columns.str.lower().str.replace(' ', '_')
    df.columns = df.columns.str.replace('.', '_')

    # Remove any duplicate columns
    df = df.loc[:, ~df.columns.duplicated()]

    # Get the maximum length of each column
    column_lengths = df.astype(str).apply(lambda x: x.str.len()).max().tolist()

    # Limit column names to 64 characters
    df.columns = df.columns.str[:64]

    # Replace NaN values with empty string
    df = df.replace({np.nan: ''})

    # Add the 'uploaddate' column with the format 'yyyymmdd'
    dateObj = pd.to_datetime('today').strftime("%Y%m%d")
    df['uploaddate'] = dateObj


   
    # Create the stockinfo table in MySQL
    cursor = cnx.cursor()
    cursor.execute(f"DROP TABLE IF EXISTS {table_name}")
   
    create_table_query = f"CREATE TABLE {table_name} ("
    for col, length in zip(df.columns, column_lengths):
        create_table_query += f"{col} varchar({length}), "
    create_table_query += "uploaddate varchar(8))"

    create_table_query = f"CREATE TABLE {table_name} ("
    for col, length in zip(df.columns, column_lengths):
        if col in numeric_cols:
            col_data_type = 'float'
        else:
            col_data_type = f"varchar({length})"
        create_table_query += f"{col} {col_data_type}, "
    create_table_query += "uploaddate varchar(8))"
    cursor.execute(create_table_query)

   
    # Insert data into the stockinfo table
    insert_query = f"INSERT INTO {table_name} ({', '.join(df.columns)}) VALUES ({', '.join(['%s'] * len(df.columns))})"
    values = df.values.tolist()
    cursor.executemany(insert_query, values)

    # Commit changes and close the connection
    cnx.commit()
    cursor.close()

    print(f"Stockinfo table '{table_name}' created successfully.")

def func4(csv_file, table_name,cnx):
    # Read the CSV file
    df = pd.read_csv(csv_file)
    print('read csv')

    # Convert column names to lowercase and replace spaces with underscores
    df.columns = df.columns.str.lower().str.replace(' ', '_')
    df.columns = df.columns.str.replace('.', '_')

    # Remove columns ending with _1, _2, _3, _4, _5, _6, _7, _8
    df = df.loc[:, ~df.columns.str.endswith(('_1', '_2', '_3', '_4', '_5', '_6', '_7', '_8'))]

    # Remove any duplicate columns
    df = df.loc[:, ~df.columns.duplicated()]

    # Get the maximum length of each column
    column_lengths = df.astype(str).apply(lambda x: x.str.len()).max().tolist()

    # Limit column names to 64 characters
    df.columns = df.columns.str[:64]

    # Convert numeric columns to float and replace empty or null values with NaN
    numeric_cols = df.select_dtypes(include=[np.number]).columns
    df[numeric_cols] = df[numeric_cols].apply(pd.to_numeric, errors='coerce')

    # Add the 'uploaddate' column with the format 'yyyymmdd'
    dateObj = pd.to_datetime('today').strftime("%Y%m%d")
    df['uploaddate'] = dateObj  # Add the uploaddate column with the same date for all rows

    # Create the stockinfo table in MySQL
    cursor = cnx.cursor()
    cursor.execute(f"DROP TABLE IF EXISTS {table_name}")

    create_table_query = f"CREATE TABLE {table_name} ("
    for col, length in zip(df.columns, column_lengths):
        col_data_type = 'FLOAT' if col in numeric_cols else f"VARCHAR({length})"
        create_table_query += f"{col} {col_data_type}, "
    create_table_query += "uploaddate VARCHAR(8))"

    cursor.execute(create_table_query)

    # Insert data into the stockinfo table
    for row in df.itertuples(index=False, name=None):
        row_without_nan = [value if pd.notna(value) else None for value in row]
        insert_query = f"INSERT INTO {table_name} VALUES ({', '.join(['%s'] * len(row_without_nan))})"
        cursor.execute(insert_query, row_without_nan)

    # Commit changes and close the connection
    cnx.commit()
    cursor.close()

    print(f"Stockinfo table '{table_name}' created successfully.")


def func16(table_name,cnx):
    try:
        engine1 = create_engine(f'mysql+mysqlconnector://{db_user}:{db_password}@{ip_addr}')

        # Read account_data table into holding_data DataFrame
        holding_data = pd.read_sql_table(f'{os.getenv("HOLDING_TABLE")}', con=engine1, schema=f'{os.getenv("DB_DATABASE1")}')

        df = pd.DataFrame(columns=['instrument', 'holding_summary'])

        for index, row in holding_data.iterrows():
            instr = row['instrument']
            acc_id = row['account_id']
            qty = row['qty_']

            if instr not in df['instrument'].values:
                df.loc[len(df)] = [instr, f'{acc_id}:{qty} (Total Qty: {qty})']
            else:
                idx = df.index[df['instrument'] == instr].tolist()[0]
                current_summary = df.loc[idx, 'holding_summary']
                #To extract the total quantity from the current summary
                current_qty = float(current_summary.split(' (Total Qty: ')[-1][:-1])
                # Update the total quantity
                total_qty = current_qty + float(qty)
                # Add the new account_id and qty to the holding_summary
                new_summary = f"{current_summary.split(' (Total Qty: ')[0]}, {acc_id}:{qty} (Total Qty: {total_qty})"
                df.loc[idx, 'holding_summary'] = new_summary


        print("Total number of distinct instruments in holdings data: ",df.shape[0])

        # Get the maximum length of each column
        column_lengths = df.astype(str).apply(lambda x: x.str.len()).max().tolist()

        # Limit column names to 64 characters
        df.columns = df.columns.str[:64]

        # Convert numeric columns to float and replace empty or null values with NaN
        numeric_cols = df.select_dtypes(include=[np.number]).columns
        df[numeric_cols] = df[numeric_cols].apply(pd.to_numeric, errors='coerce')

        #Create table in MySQL
        cursor = cnx.cursor()
        cursor.execute(f"DROP TABLE IF EXISTS {table_name}")
        create_table_query = f"CREATE TABLE {table_name} ("
        for col, length in zip(df.columns, column_lengths):
            col_data_type = 'FLOAT' if col in numeric_cols else f"VARCHAR({length})"
            create_table_query += f"{col} {col_data_type}, "
        create_table_query = create_table_query.rstrip(", ") + ")"
        cursor.execute(create_table_query)

        # Insert data into the stockinfo table
        for row in df.itertuples(index=False, name=None):
            row_without_nan = [value if pd.notna(value) else None for value in row]
            insert_query = f"INSERT INTO {table_name} VALUES ({', '.join(['%s'] * len(row_without_nan))})"
            cursor.execute(insert_query, row_without_nan)

        # Commit changes and close the connection
        cnx.commit()
        cursor.close()

        print(f"'{table_name}' table created successfully.")

    except Exception as e:
        print(f"Error: {e}")




#In destination_table: Add holding_summary column manually.
# ALTER TABLE alldata.stockinfo
# ADD holding_summary VARCHAR(100);
#func19: create index on nse_code column for #first time# to update holding_summary column fast.

def func19(table_data, destination_table, cnx,uploaddate):
    try:
        engine = create_engine(f'mysql+mysqlconnector://{db_user}:{db_password}@{ip_addr}/{db_database}?connect_timeout=60')

        # Read data from the account_summary table into a DataFrame
        query = f"SELECT * FROM {table_data}"
        df = pd.read_sql(query, con=engine)

        sql_query = f"SELECT * FROM {destination_table} WHERE uploaddate = '{uploaddate}'"
        stock_data = pd.read_sql(sql_query, con=engine)
        print(stock_data.shape[0])
        
        #create index on nse_code column for first time
        cursor = cnx.cursor()

        # query = f"CREATE INDEX idx_nse_code ON {destination_table} (nse_code)"
        
        # cursor.execute(query)

        # Prepare data for batch update
        update_data = []
        for index, row in df.iterrows():
            instr = row['instrument']
            holding_summ = row['holding_summary']
            update_data.append((holding_summ, instr))

        # Update holding_summary for matching nse_code
        update_query = f"""
        UPDATE {destination_table} 
        SET holding_summary = %s 
        WHERE nse_code = %s AND uploaddate = '{uploaddate}'
        """
        cursor.executemany(update_query, update_data)

        # Commit changes and close the connection
        cnx.commit()
        cursor.close()
        cnx.close()

        print(f"{destination_table} updated.")

    except Exception as e:
        print(f"Error: {e}")


import logging
import sys
# Set up logging configuration to print logs to the screen
logging.basicConfig(level=logging.DEBUG, format='%(levelname)s: %(message)s', stream=sys.stdout)


def func21(commented_path, destination_table,uploaddate):
    try:
        # Load the dataframes
        df = pd.read_excel(commented_path)

        # Load the workbook and select the active worksheet
        wb = load_workbook(commented_path)
        ws = wb.active

        # Define the fill for highlighting
        highlight_fill = PatternFill(start_color="FFD580", end_color="FFD580",  fill_type="solid")

        # Create a connection to the MySQL database
        engine = create_engine(f'mysql+mysqlconnector://{db_user}:{db_password}@{ip_addr}/  {db_database}?connect_timeout=60')
    
        # Load the stock data from the MySQL table
        stock_data_query = f"SELECT name, holding_summary FROM {destination_table} WHERE    uploaddate = '{uploaddate}'"
        stock_data = pd.read_sql(stock_data_query, con=engine)

        # Convert stock_data to a dictionary for quick lookup
        stock_data_dict = stock_data.set_index('name')['holding_summary'].to_dict()

        for index, row in df.iterrows():
            name = row['name']
            if name in stock_data_dict:
                holding_summary = stock_data_dict[name]

                # Check if holding_summary is not empty
                if pd.notna(holding_summary) and holding_summary != "":
                    # Highlight the corresponding cell in the second column
                    cell_to_highlight = ws.cell(row=index + 2, column=2)  # since header is in the first row
                    cell_to_highlight.fill = highlight_fill

                    # Add comment to the corresponding cell in the second column
                    cell_to_comment = ws.cell(row=index + 2, column=2)
                    cell_to_comment.comment = Comment(text=str(holding_summary), author="")

        
        # Save the updated Excel file
        output_filename = f'holding_highlighted_comment_bank_{uploaddate}.xlsx'
        wb.save(output_filename)

        # Email configuration
        smtp_host = os.getenv("SMTP_HOST")
        smtp_port = os.getenv("SMTP_PORT")
        sender_email = os.getenv("SENDER_EMAIL")
        sender_app_password = os.getenv("SENDER_APP_PASSWORD")  # You need to generate an App Password in your Gmail account settings
        recipient_email = os.getenv("RECIPIENT_EMAIL")
        

        # Create the email
        msg = MIMEMultipart()
        msg['From'] = sender_email
        msg['To'] = recipient_email
        msg['Subject'] = 'Holdings highlighted and commented bank data'

        # Attach the Excel file to the email
        with open(output_filename, "rb") as excel_file:
            part = MIMEApplication(excel_file.read(), Name=output_filename)
            part['Content-Disposition'] = f'attachment; filename="{output_filename}"'
            msg.attach(part)

        # Connect to the Gmail SMTP server and send the email
        with smtplib.SMTP(smtp_host, smtp_port) as server:
            server.starttls()
            server.login(sender_email, sender_app_password)
            server.sendmail(sender_email, recipient_email, msg.as_string())

            print("func21 executed successfully.")

    except Exception as e:
        print(f"Error: {e}")
        


def func22(commented_nonbank_path, destination_table,uploaddate):
    try:
        # Load the dataframes
        df = pd.read_excel(commented_nonbank_path)

        # Load the workbook and select the active worksheet
        wb = load_workbook(commented_nonbank_path)
        ws = wb.active

        # Define the fill for highlighting
        highlight_fill = PatternFill(start_color="FFD580", end_color="FFD580",  fill_type="solid")

        # Create a connection to the MySQL database
        engine = create_engine(f'mysql+mysqlconnector://{db_user}:{db_password}@{ip_addr}/  {db_database}?connect_timeout=60')
    
        # Load the stock data from the MySQL table
        stock_data_query = f"SELECT name, holding_summary FROM {destination_table} WHERE    uploaddate = '{uploaddate}'"
        stock_data = pd.read_sql(stock_data_query, con=engine)

        # Convert stock_data to a dictionary for quick lookup
        stock_data_dict = stock_data.set_index('name')['holding_summary'].to_dict()

        for index, row in df.iterrows():
            name = row['name']
            if name in stock_data_dict:
                holding_summary = stock_data_dict[name]

                # Check if holding_summary is not empty
                if pd.notna(holding_summary) and holding_summary != "":
                    # Highlight the corresponding cell in the second column
                    cell_to_highlight = ws.cell(row=index + 2, column=2)  # Assuming header is in the first row
                    cell_to_highlight.fill = highlight_fill

                    # Add comment to the corresponding cell in the second column
                    cell_to_comment = ws.cell(row=index + 2, column=2)
                    cell_to_comment.comment = Comment(text=str(holding_summary), author="")

        # Save the updated Excel file
        output_filename = f'holding_highlighted_comment_nonbank_{uploaddate}.xlsx'
        wb.save(output_filename)

        # Email configuration
        smtp_host = os.getenv("SMTP_HOST")
        smtp_port = os.getenv("SMTP_PORT")
        sender_email = os.getenv("SENDER_EMAIL")
        sender_app_password = os.getenv("SENDER_APP_PASSWORD")  # You need to generate anApp Password in your Gmail account settings
        recipient_email = os.getenv("RECIPIENT_EMAIL")
        

        # Create the email
        msg = MIMEMultipart()
        msg['From'] = sender_email
        msg['To'] = recipient_email
        msg['Subject'] = 'Holdings highlighted and commented non-bank data'

        # Attach the Excel file to the email
        with open(output_filename, "rb") as excel_file:
            part = MIMEApplication(excel_file.read(), Name=output_filename)
            part['Content-Disposition'] = f'attachment; filename="{output_filename}"'
            msg.attach(part)

        # Connect to the Gmail SMTP server and send the email
        with smtplib.SMTP(smtp_host, smtp_port) as server:
            server.starttls()
            server.login(sender_email, sender_app_password)
            server.sendmail(sender_email, recipient_email, msg.as_string())

            print("func22 executed successfully.")

    except Exception as e:
        print(f"Error: {e}") 


def func23(csv_file, table_name, uploaddate, cnx):
    try:
        # Read the CSV file
        df = pd.read_csv(csv_file)
        print('Read CSV')

        # Convert column names to lowercase and replace spaces with underscores
        df.columns = df.columns.str.lower().str.replace(' ', '_')
        df.columns = df.columns.str.replace('.', '_')

        # Remove columns ending with _1, _2, _3, _4, _5, _6, _7, _8
        df = df.loc[:, ~df.columns.str.endswith(('_1', '_2', '_3', '_4', '_5', '_6', '_7', '_8'))]

        # Remove any duplicate columns
        df = df.loc[:, ~df.columns.duplicated()]

        # Get the maximum length of each column
        column_lengths = df.astype(str).apply(lambda x: x.str.len()).max().tolist()

        # Limit column names to 64 characters
        df.columns = df.columns.str[:64]

        # Convert numeric columns to float and replace empty or null values with NaN
        numeric_cols = df.select_dtypes(include=[np.number]).columns
        df[numeric_cols] = df[numeric_cols].apply(pd.to_numeric, errors='coerce')

    
        print(f'Number of columns in csv file: {df.shape[1]}')

        # Create the table in MySQL
        cursor = cnx.cursor()
        cursor.execute(f"DROP TABLE IF EXISTS {table_name}")

        create_table_query = f"CREATE TABLE {table_name} ("
        for col, length in zip(df.columns, column_lengths):
            col_data_type = 'FLOAT' if col in numeric_cols else f"VARCHAR({length})"
            create_table_query += f"{col} {col_data_type}, "
        create_table_query = create_table_query.rstrip(", ") + ")"

        cursor.execute(create_table_query)

        # Insert data into the table
        for row in df.itertuples(index=False, name=None):
            row_without_nan = [value if pd.notna(value) else None for value in row]
            insert_query = f"INSERT INTO {table_name} VALUES ({', '.join(['%s'] * len(row_without_nan))})"
            cursor.execute(insert_query, row_without_nan)

        # Commit changes and close the connection
        cnx.commit()
        cursor.close()

        print(f"Table '{table_name}' created and data inserted successfully.")
    except Exception as e:
        print(f"Error: {e}")
 
#Create new table in mysql nse_bse_table:
def func17(csv_file, table_name,cnx):
    try:
        # Read the CSV file
        df = pd.read_csv(csv_file)
        print('read csv')

        # Convert column names to lowercase and replace spaces with underscores
        df.columns = df.columns.str.lower().str.replace(' ', '_')
        df.columns = df.columns.str.replace('.', '_')

        # Remove columns ending with _1, _2, _3, _4, _5, _6, _7, _8
        df = df.loc[:, ~df.columns.str.endswith(('_1', '_2', '_3', '_4', '_5', '_6', '_7',  '_8'))]

        # Remove any duplicate columns
        df = df.loc[:, ~df.columns.duplicated()]

        # Get the maximum length of each column
        column_lengths = df.astype(str).apply(lambda x: x.str.len()).max().tolist()

        # Limit column names to 64 characters
        df.columns = df.columns.str[:64]

        # Convert numeric columns to float and replace empty or null values with NaN
        numeric_cols = df.select_dtypes(include=[np.number]).columns
        df[numeric_cols] = df[numeric_cols].apply(pd.to_numeric, errors='coerce')

        # Create the stockinfo table in MySQL
        cursor = cnx.cursor()
        cursor.execute(f"DROP TABLE IF EXISTS {table_name}")

        create_table_query = f"CREATE TABLE {table_name} ("
        for col, length in zip(df.columns, column_lengths):
            col_data_type = 'FLOAT' if col in numeric_cols else f"VARCHAR({length})"
            create_table_query += f"{col} {col_data_type}, "
        create_table_query = create_table_query.rstrip(", ") + ")"

        cursor.execute(create_table_query)

        # Insert data into the stockinfo table
        for row in df.itertuples(index=False, name=None):
            row_without_nan = [value if pd.notna(value) else None for value in row]
            insert_query = f"INSERT INTO {table_name} VALUES ({', '.join(['%s'] * len   (row_without_nan))})"
            cursor.execute(insert_query, row_without_nan)

        # Commit changes and close the connection
        cnx.commit()
        cursor.close()

        print(f"'{table_name}' table created successfully.")
    
    except Exception as e:
        print(f"Error: {e}")

#Create an index on the name column to speed up the update process only for first time.
def func18(table1, table2, cnx,uploaddate):
    try:
        engine = create_engine(f'mysql+mysqlconnector://{db_user}:{db_password}@{ip_addr}/  {db_database}?connect_timeout=60')

        query = f"SELECT * FROM {table1} WHERE uploaddate = '{uploaddate}'"
        data1 = pd.read_sql(query,con = engine)

        sql_query = f"SELECT * FROM {table2}"
        data2 = pd.read_sql(sql_query,con = engine)

        for i in data2['bse_code']:
            #print(i)
            for j in range(len(data1['name'])):
                if i == data1['bse_code'][j]:
                    data1.at[j, 'nse_code'] = data2[data2['bse_code'] == i]['instrument'].  values 


        for instrument_entry in data2['instrument']:
            part_before_dash = instrument_entry.split('-')[0]
            mask = data1['nse_code'] == part_before_dash
            if any(mask):
                zerodha_value = data2.loc[data2['instrument'] == instrument_entry,  'instrument'].values[0]
                # print(f"Match found for {part_before_dash}. Updating with {zerodha_value}")
                data1.loc[mask, 'nse_code'] = zerodha_value

        cursor = cnx.cursor()

        # Create an index on the name column to speed up the update process
        # cursor.execute(f"CREATE INDEX idx_name ON {table1}(name)")

        # Modify the nse_code column to increase its length
        cursor.execute(f"ALTER TABLE {table1} MODIFY COLUMN nse_code VARCHAR(100)")

        for index, row in data1.iterrows():
                name = row['name']
                nse_code = row['nse_code']
                if pd.notna(nse_code):  # Update only if nse_code is not NaN
                    update_query = f'UPDATE {table1} SET nse_code = "{nse_code}" WHERE name = "{name}" AND uploaddate = "{uploaddate}"'
                    cursor.execute(update_query)
                    # print(f"Updated nse_code for {name} to {nse_code}")
        cnx.commit()
        cursor.close()

        print(f"'{table1}' table updated.")
    
    except Exception as e:
        print(f"Error: {e}")


def func20(table1,destination_table,cnx,uploaddate):
    try:
        engine = create_engine(f'mysql+mysqlconnector://{db_user}:{db_password}@{ip_addr}/  {db_database}?connect_timeout=60')

        query = f'SELECT * from {table1}'
        data1 = pd.read_sql(query, con = engine)

        sql_query = f'SELECT * FROM {destination_table} WHERE uploaddate = "{uploaddate}"'
        data2 = pd.read_sql(sql_query, con = engine)

        # Define the SQL query to extract the required list
        extraction_query = f"""
        SELECT a.*
        FROM {table1} a
        LEFT JOIN {destination_table} s
        ON a.instrument = s.nse_code 
        AND s.uploaddate = '{uploaddate}'
        WHERE s.nse_code IS NULL;
        """
        # Execute the extraction query and load the data into a DataFrame
        extracted_data = pd.read_sql(extraction_query, con=engine)

        # Save the extracted data to a CSV file (optional)
        extracted_data.to_excel("extracted_data.xlsx", index=False)
        print(f"CSV file created successfully...")

        # Construct the output Excel file name with the upload date
        output_filename = f'list_of_stocks_not_in_nse_code.xlsx'
        extracted_data.to_excel(output_filename, index=False)

        # Email configuration
        smtp_host = os.getenv("SMTP_HOST")
        smtp_port = os.getenv("SMTP_PORT")
        sender_email = os.getenv("SENDER_EMAIL")
        sender_app_password = os.getenv("SENDER_APP_PASSWORD")  # You need to generate an App Password in your Gmail account settings
        recipient_email = os.getenv("RECIPIENT_EMAIL")
        

        # Create the email
        msg = MIMEMultipart()
        msg['From'] = sender_email
        msg['To'] = recipient_email
        msg['Subject'] = 'List of stocks absent in nse_code column but present in holding data'

        # Attach the Excel file to the email
        with open(output_filename, "rb") as excel_file:
            part = MIMEApplication(excel_file.read(), Name=output_filename)
            part['Content-Disposition'] = f'attachment; filename="{output_filename}"'
            msg.attach(part)

        # Connect to the Gmail SMTP server and send the email
        with smtplib.SMTP(smtp_host, smtp_port) as server:
            server.starttls()
            server.login(sender_email, sender_app_password)
            server.sendmail(sender_email, recipient_email, msg.as_string())
        
        print(f"{output_filename} sent successfully...")

    except Exception as e:
        print(f"Error: {e}")



def func24(holding_file, destination_table, cnx, uploaddate):
    try:
        engine = create_engine(f'mysql+mysqlconnector://{db_user}:{db_password}@{ip_addr}/{db_database}?connect_timeout=60')

        query = f"SELECT * FROM {destination_table} WHERE uploaddate = '{uploaddate}'"
        data1 = pd.read_sql(query, con=engine)
        
        data2 = pd.read_excel(holding_file)
        
        wb = load_workbook(holding_file)
        ws = wb.active

        # Define the fill for highlighting
        highlight_fill = PatternFill(start_color="f4c2c2", end_color="f4c2c2", fill_type="solid")

        for index, row in data1.iterrows():
            name = row['name']
            volume = row['volume']
            volume_1week = row['volume_1week_average']
            volume_1month = row['volume_1month_average']
            m_cap = row['market_capitalization']
            if volume_1week != 0 and volume >= volume_1week * 1.5:
                # Find the corresponding row in the Excel file
                for excel_row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    excel_name_cell = excel_row[0]  # since 'name' is in the first column
                    if excel_name_cell.value == name:
                        # To highlight the market capitalization cell
                        excel_m_cap_cell = excel_row[2]  # Since 'market_capitalization' is in the third column
                        excel_m_cap_cell.fill = highlight_fill

                        # Add a comment to the cell
                        comment_text = (f"Volume: {volume}\n"
                                        f"% chg: {volume / volume_1week:.2f}x\n"
                                        f"1week vol.: {volume_1week}\n"
                                        f"1month vol.: {volume_1month}")
                        comment = Comment(comment_text, "Author")
                        excel_m_cap_cell.comment = comment
                        break
        

        # Save the updated Excel file
        output_filename = f'Volume_commented_bank_{uploaddate}.xlsx'
        wb.save(output_filename)

        # Email configuration
        smtp_host = os.getenv("SMTP_HOST")
        smtp_port = os.getenv("SMTP_PORT")
        sender_email = os.getenv("SENDER_EMAIL")
        sender_app_password = os.getenv("SENDER_APP_PASSWORD")  # You need to generate an App Password in your Gmail account settings
        recipient_email = os.getenv("RECIPIENT_EMAIL")
        

        # Create the email
        msg = MIMEMultipart()
        msg['From'] = sender_email
        msg['To'] = recipient_email
        msg['Subject'] = 'Commented and highlighted bank volume'

        # Attach the Excel file to the email
        with open(output_filename, "rb") as excel_file:
            part = MIMEApplication(excel_file.read(), Name=output_filename)
            part['Content-Disposition'] = f'attachment; filename="{output_filename}"'
            msg.attach(part)

        # Connect to the Gmail SMTP server and send the email
        with smtplib.SMTP(smtp_host, smtp_port) as server:
            server.starttls()
            server.login(sender_email, sender_app_password)
            server.sendmail(sender_email, recipient_email, msg.as_string())
        
        print(f"{output_filename} sent successfully...")


    except Exception as e:
        print(f"Error: {e}")


def func25(holding_file, destination_table, cnx, uploaddate):
    try:
        engine = create_engine(f'mysql+mysqlconnector://{db_user}:{db_password}@{ip_addr}/{db_database}?connect_timeout=60')

        query = f"SELECT * FROM {destination_table} WHERE uploaddate = '{uploaddate}'"
        data1 = pd.read_sql(query, con=engine)
        
        data2 = pd.read_excel(holding_file)
        
        wb = load_workbook(holding_file)
        ws = wb.active

        # Define the fill for highlighting
        highlight_fill = PatternFill(start_color="f4c2c2", end_color="f4c2c2", fill_type="solid")

        for index, row in data1.iterrows():
            name = row['name']
            volume = row['volume']
            volume_1week = row['volume_1week_average']
            volume_1month = row['volume_1month_average']
            m_cap = row['market_capitalization']
            if volume_1week != 0 and volume >= volume_1week * 1.5:
                # Find the corresponding row in the Excel file
                for excel_row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    excel_name_cell = excel_row[0]  # since 'name' is in the first column
                    if excel_name_cell.value == name:
                        # Highlight the market capitalization cell
                        excel_m_cap_cell = excel_row[2]  # Since 'market_capitalization' is in the third column
                        excel_m_cap_cell.fill = highlight_fill

                        # Add a comment to the cell
                        comment_text = (f"Volume: {volume}\n"
                                        f"% chg: {volume / volume_1week:.2f}x\n"
                                        f"1week vol.: {volume_1week}\n"
                                        f"1month vol.: {volume_1month}")
                        comment = Comment(comment_text, "Author")
                        excel_m_cap_cell.comment = comment
                        break

        # Save the updated Excel file
        output_filename = f'Volume_commented_nonbank_{uploaddate}.xlsx'
        wb.save(output_filename)

        # Email configuration
        smtp_host = os.getenv("SMTP_HOST")
        smtp_port = os.getenv("SMTP_PORT")
        sender_email = os.getenv("SENDER_EMAIL")
        sender_app_password = os.getenv("SENDER_APP_PASSWORD")  # You need to generate an App Password in your Gmail account settings
        recipient_email = os.getenv("RECIPIENT_EMAIL")
        

        # Create the email
        msg = MIMEMultipart()
        msg['From'] = sender_email
        msg['To'] = recipient_email
        msg['Subject'] = 'Commented and highlighted non-bank volume'

        # Attach the Excel file to the email
        with open(output_filename, "rb") as excel_file:
            part = MIMEApplication(excel_file.read(), Name=output_filename)
            part['Content-Disposition'] = f'attachment; filename="{output_filename}"'
            msg.attach(part)

        # Connect to the Gmail SMTP server and send the email
        with smtplib.SMTP(smtp_host, smtp_port) as server:
            server.starttls()
            server.login(sender_email, sender_app_password)
            server.sendmail(sender_email, recipient_email, msg.as_string())
        
        print(f"{output_filename} sent successfully...")


    except Exception as e:
        print(f"Error: {e}")


# fnotable created:
def func26(file_path, table_name, cnx):
    try:
        df = pd.read_excel(file_path)
        print("Frist data:\n",df.head())
        df.columns = df.columns.str.lower().str.replace(' ', '_').str.replace('%', 'perct').str.replace('.', '_')

        # Get the maximum length of each column
        column_lengths = df.astype(str).apply(lambda x: x.str.len()).max().tolist()

        # Limit column names to 64 characters
        df.columns = df.columns.str[:64]
    
        # Convert numeric columns to float and replace empty or null values with NaN
        numeric_cols = df.select_dtypes(include=[np.number]).columns
        df[numeric_cols] = df[numeric_cols].apply(pd.to_numeric, errors='coerce')

        # Create table in MySQL
        cursor = cnx.cursor()
        cursor.execute(f"DROP TABLE IF EXISTS {table_name}")
        create_table_query = f"CREATE TABLE {table_name} ("
        for col, length in zip(df.columns, column_lengths):
            col_data_type = 'DOUBLE' if col in numeric_cols else f"VARCHAR({length})"
            create_table_query += f"{col} {col_data_type}, "
        create_table_query = create_table_query.rstrip(", ") + ")"
        
        cursor.execute(create_table_query)

        # Insert data into the table
        for row in df.itertuples(index=False, name=None):
            row_without_nan = [value if pd.notna(value) else None for value in row]
            insert_query = f"INSERT INTO {table_name} VALUES ({', '.join(['%s'] * len(row_without_nan))})"
            cursor.execute(insert_query, row_without_nan)

        # Commit changes and close the connection
        cnx.commit()
        cursor.close()

        print(f"'{table_name}' table created successfully.")
        print("func26 executed successfully...")

    except Exception as e:
        print(f"Error: {e}")


# ADD FNO COLUMN
def func27(table, destination_table, file_path, cnx, uploaddate):
    try:
        # Create the SQLAlchemy engine
        engine = create_engine(f'mysql+mysqlconnector://{db_user}:{db_password}@{ip_addr}/{db_database}?connect_timeout=60')

        # Read the data from the SQL tables
        sql_query = f'SELECT * FROM {table}'
        data1 = pd.read_sql(sql_query, con=engine)
        print("Columns of data1 are:", data1.columns)

        query = f"SELECT * FROM {destination_table} WHERE uploaddate = '{uploaddate}'"
        data2 = pd.read_sql(query, con=engine)

        # Load the Excel file
        wb = load_workbook(file_path)
        ws = wb.active

        # Add a new column 'fno' to the Excel file
        ws['BB1'] = 'fno'  # since the 'fno' column should be in column BB

        # Create dictionaries for faster lookups
        data1_symbols = set(data1['symbol'].dropna())
        data2_dict = {row['name']: row['nse_code'] for _, row in data2.iterrows()}

        # Collect updates in a list to batch update the Excel file
        updates = []

        for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1, values_only=True), start=2):
            name = row[0]
            matched = False
            if name in data2_dict:
                nse_code = data2_dict[name]
                if nse_code in data1_symbols:
                    updates.append((idx, 'Y'))
                    matched = True
            if not matched:
                updates.append((idx, 'N'))

        # Batch update the Excel file
        for idx, value in updates:
            ws.cell(row=idx, column=54, value=value)  # since 'fno' column is column BB

        # Save the updated Excel file
        output_filename = 'volume_fno_commented_bank.xlsx'
        wb.save(output_filename)

        # Email configuration
        smtp_host = os.getenv("SMTP_HOST")
        smtp_port = os.getenv("SMTP_PORT")
        sender_email = os.getenv("SENDER_EMAIL")
        sender_app_password = os.getenv("SENDER_APP_PASSWORD")  # You need to generate an App Password in your Gmail account settings
        recipient_email = os.getenv("RECIPIENT_EMAIL")
        

        # Create the email
        msg = MIMEMultipart()
        msg['From'] = sender_email
        msg['To'] = recipient_email
        msg['Subject'] = 'Commented and added fno column bank volume'

        # Attach the Excel file to the email
        with open(output_filename, "rb") as excel_file:
            part = MIMEApplication(excel_file.read(), Name=output_filename)
            part['Content-Disposition'] = f'attachment; filename="{output_filename}"'
            msg.attach(part)

        # Connect to the Gmail SMTP server and send the email
        with smtplib.SMTP(smtp_host, smtp_port) as server:
            server.starttls()
            server.login(sender_email, sender_app_password)
            server.sendmail(sender_email, recipient_email, msg.as_string())
        
        print(f"{output_filename} sent successfully...")

        print("func27 executed successfully.")

    except Exception as e:
        print(f"Error: {e}")



def func28(table, destination_table, file_path, cnx, uploaddate):
    try:
        # Create the SQLAlchemy engine
        engine = create_engine(f'mysql+mysqlconnector://{db_user}:{db_password}@{ip_addr}/{db_database}?connect_timeout=60')

        # Read the data from the SQL tables
        sql_query = f'SELECT * FROM {table}'
        data1 = pd.read_sql(sql_query, con=engine)
        print("Columns of data1 are:", data1.columns)

        query = f"SELECT * FROM {destination_table} WHERE uploaddate = '{uploaddate}'"
        data2 = pd.read_sql(query, con=engine)

        # Load the Excel file
        wb = load_workbook(file_path)
        ws = wb.active

        # Add a new column 'fno' to the Excel file
        ws['BD1'] = 'fno'  # since the 'fno' column should be in column BD

        # Create dictionaries for faster lookups
        data1_symbols = set(data1['symbol'].dropna())
        data2_dict = {row['name']: row['nse_code'] for _, row in data2.iterrows()}

        # Collect updates in a list to batch update the Excel file
        updates = []

        for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1, values_only=True), start=2):
            name = row[0]
            matched = False
            if name in data2_dict:
                nse_code = data2_dict[name]
                if nse_code in data1_symbols:
                    updates.append((idx, 'Y'))
                    matched = True
            if not matched:
                updates.append((idx, 'N'))

        # Batch update the Excel file
        for idx, value in updates:
            ws.cell(row=idx, column=56, value=value)  # since 'fno' column is column BD

        # Save the updated Excel file
        output_filename = 'volume_fno_commented_nonbank.xlsx'
        wb.save(output_filename)

        # Email configuration
        smtp_host = os.getenv("SMTP_HOST")
        smtp_port = os.getenv("SMTP_PORT")
        sender_email = os.getenv("SENDER_EMAIL")
        sender_app_password = os.getenv("SENDER_APP_PASSWORD")  # You need to generate an App Password in your Gmail account settings
        recipient_email = os.getenv("RECIPIENT_EMAIL")
        

        # Create the email
        msg = MIMEMultipart()
        msg['From'] = sender_email
        msg['To'] = recipient_email
        msg['Subject'] = 'Commented and added fno column non-bank volume'

        # Attach the Excel file to the email
        with open(output_filename, "rb") as excel_file:
            part = MIMEApplication(excel_file.read(), Name=output_filename)
            part['Content-Disposition'] = f'attachment; filename="{output_filename}"'
            msg.attach(part)

        # Connect to the Gmail SMTP server and send the email
        with smtplib.SMTP(smtp_host, smtp_port) as server:
            server.starttls()
            server.login(sender_email, sender_app_password)
            server.sendmail(sender_email, recipient_email, msg.as_string())
        
        print(f"{output_filename} sent successfully...")

        print("func28 executed successfully.")

    except Exception as e:
        print(f"Error: {e}")



# def func29(file_path,cnx):
#     try:
#         cursor = cnx.cursor()
#         query = f"SELECT codedesc1 FROM misccodes WHERE codetype = 'yoy_quarterly_sales_growth'"

#         cursor.execute(query)
#         result = cursor.fetchall()
#         print(result)
#         cursor.close()
#         cnx.close()
#         color1 = result[0][0]
#         print("Color1 is:",color1)
#         color2 = result[1][0]
#         color3 = result[2][0]
#         color4 = result[3][0]
#         color5 = result[4][0]
        
#         # Load the Excel file
#         wb = load_workbook(file_path)
#         sheet = wb.active
#         col1_index = 4
#         col2_index = 31

#         # Swap column names
#         sheet.cell(row=1, column=col1_index).value, sheet.cell(row=1, column=col2_index). value = sheet.cell(row=1, column=col2_index).value, sheet.cell(row=1,  column=col1_index).value

#         # Iterate over the rows and swap the values
#         for row in sheet.iter_rows(min_row=2):  # Since the data starts from row 2
#             # Swap the values in the specified columns
#             col1_value = row[col1_index - 1].value
#             col2_value = row[col2_index - 1].value
#             row[col1_index - 1].value = col2_value
#             row[col2_index - 1].value = col1_value
            
#             col1_comment = row[col1_index - 1].comment
#             col2_comment = row[col2_index - 1].comment
#             row[col1_index - 1].comment = col2_comment
#             row[col2_index - 1].comment = col1_comment

#             # Preserve cell styles
#             col1_cell = row[col1_index - 1]
#             col2_cell = row[col2_index - 1]
#             col1_cell.font, col2_cell.font = Font(**col2_cell.font.__dict__), Font  (**col1_cell.font.__dict__)
#             col1_cell.fill, col2_cell.fill = PatternFill(**col2_cell.fill.__dict__),    PatternFill(**col1_cell.fill.__dict__)
#             col1_cell.border, col2_cell.border = Border(**col2_cell.border.__dict__), Border(**col1_cell.border.__dict__)

#             # Apply grey shading based on yoy_qyarterly_sales_growth values:
#             yoy_value = col2_value
#             #By default fill white color
#             fill = PatternFill(start_color=color1, end_color=color1, fill_type="solid")

#             if yoy_value is not None and isinstance(yoy_value, (int, float)):
#                 if 10 <= yoy_value < 15:
#                     fill = PatternFill(start_color=color2, end_color=color2, fill_type="solid")  
#                 elif 15 <= yoy_value < 20:
#                     fill = PatternFill(start_color=color3, end_color=color3, fill_type="solid")  
#                 elif 20 <= yoy_value < 25:
#                     fill = PatternFill(start_color=color4, end_color=color4, fill_type="solid")  
#                 elif yoy_value >= 25:
#                     fill = PatternFill(start_color=color5, end_color=color5, fill_type="solid")  

#             col1_cell.fill = fill


#         # Save the modified Excel file
#         output_filename = 'quarterly_sales_highlight_bank11.xlsx'
#         wb.save(output_filename)
#         print("File saved successfully.")

#         # # Email configuration
#         # smtp_host = os.getenv("SMTP_HOST")
#         # smtp_port = os.getenv("SMTP_PORT")
#         # sender_email = os.getenv("SENDER_EMAIL")
#         # sender_app_password = os.getenv("SENDER_APP_PASSWORD")  # You need to generate an App Password in your Gmail account settings
#         # recipient_email = os.getenv("RECIPIENT_EMAIL")
        

#         # # Create the email
#         # msg = MIMEMultipart()
#         # msg['From'] = sender_email
#         # msg['To'] = recipient_email
#         # msg['Subject'] = 'Highlighted yoy_quarterly_sales_growth column in Bank report'

#         # # Attach the Excel file to the email
#         # with open(output_filename, "rb") as excel_file:
#         #     part = MIMEApplication(excel_file.read(), Name=output_filename)
#         #     part['Content-Disposition'] = f'attachment; filename="{output_filename}"'
#         #     msg.attach(part)

#         # # Connect to the Gmail SMTP server and send the email
#         # with smtplib.SMTP(smtp_host, smtp_port) as server:
#         #     server.starttls()
#         #     server.login(sender_email, sender_app_password)
#         #     server.sendmail(sender_email, recipient_email, msg.as_string())
        
#         # print("func29 executed successfully...")
#     except Exception as e:
#         print(f"Error : {e}")


def func29(file_path):
    try:
        # Load the Excel file
        wb = load_workbook(file_path)
        sheet = wb.active
        col1_index = 4
        col2_index = 31

        # Swap column names
        sheet.cell(row=1, column=col1_index).value, sheet.cell(row=1, column=col2_index). value = sheet.cell(row=1, column=col2_index).value, sheet.cell(row=1,  column=col1_index).value

        # Iterate over the rows and swap the values
        for row in sheet.iter_rows(min_row=2):  # Since the data starts from row 2
            # Swap the values in the specified columns
            col1_value = row[col1_index - 1].value
            col2_value = row[col2_index - 1].value
            row[col1_index - 1].value = col2_value
            row[col2_index - 1].value = col1_value
            
            col1_comment = row[col1_index - 1].comment
            col2_comment = row[col2_index - 1].comment
            row[col1_index - 1].comment = col2_comment
            row[col2_index - 1].comment = col1_comment

            # Preserve cell styles
            col1_cell = row[col1_index - 1]
            col2_cell = row[col2_index - 1]
            col1_cell.font, col2_cell.font = Font(**col2_cell.font.__dict__), Font  (**col1_cell.font.__dict__)
            col1_cell.fill, col2_cell.fill = PatternFill(**col2_cell.fill.__dict__),    PatternFill(**col1_cell.fill.__dict__)
            col1_cell.border, col2_cell.border = Border(**col2_cell.border.__dict__), Border(**col1_cell.border.__dict__)

            # Apply grey shading based on yoy_qyarterly_sales_growth values:
            yoy_value = col2_value
            #By default fill white color
            fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

            if yoy_value is not None and isinstance(yoy_value, (int, float)):
                if 10 <= yoy_value < 15:
                    fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")  
                elif 15 <= yoy_value < 20:
                    fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")  
                elif 20 <= yoy_value < 25:
                    fill = PatternFill(start_color="B3B3B3", end_color="B3B3B3", fill_type="solid")  
                elif yoy_value >= 25:
                    fill = PatternFill(start_color="999999", end_color="999999", fill_type="solid")  

            col1_cell.fill = fill


        # Save the modified Excel file
        output_filename = 'quarterly_sales_highlight_bank.xlsx'
        wb.save(output_filename)
        print("File saved successfully.")

        # Email configuration
        smtp_host = os.getenv("SMTP_HOST")
        smtp_port = os.getenv("SMTP_PORT")
        sender_email = os.getenv("SENDER_EMAIL")
        sender_app_password = os.getenv("SENDER_APP_PASSWORD")  # You need to generate an App Password in your Gmail account settings
        recipient_email = os.getenv("RECIPIENT_EMAIL")
        

        # Create the email
        msg = MIMEMultipart()
        msg['From'] = sender_email
        msg['To'] = recipient_email
        msg['Subject'] = 'Highlighted yoy_quarterly_sales_growth column in Bank report'

        # Attach the Excel file to the email
        with open(output_filename, "rb") as excel_file:
            part = MIMEApplication(excel_file.read(), Name=output_filename)
            part['Content-Disposition'] = f'attachment; filename="{output_filename}"'
            msg.attach(part)

        # Connect to the Gmail SMTP server and send the email
        with smtplib.SMTP(smtp_host, smtp_port) as server:
            server.starttls()
            server.login(sender_email, sender_app_password)
            server.sendmail(sender_email, recipient_email, msg.as_string())
        
        print("func29 executed successfully...")
    except Exception as e:
        print(f"Error : {e}")


def func30(file_path):
    try:
        # Load the Excel file
        wb = load_workbook(file_path)
        sheet = wb.active
        col1_index = 4
        col2_index = 31

        # Swap column names
        sheet.cell(row=1, column=col1_index).value, sheet.cell(row=1, column=col2_index).   value = sheet.cell(row=1, column=col2_index).value, sheet.cell(row=1,  column=col1_index).value

        # Iterate over the rows and swap the values
        for row in sheet.iter_rows(min_row=2):  # Since the data starts from row 2
            # Swap the values in the specified columns
            col1_value = row[col1_index - 1].value
            col2_value = row[col2_index - 1].value
            row[col1_index - 1].value = col2_value
            row[col2_index - 1].value = col1_value

            
            col1_comment = row[col1_index - 1].comment
            col2_comment = row[col2_index - 1].comment
            row[col1_index - 1].comment = col2_comment
            row[col2_index - 1].comment = col1_comment

            # Preserve cell styles
            col1_cell = row[col1_index - 1]
            col2_cell = row[col2_index - 1]
            col1_cell.font, col2_cell.font = Font(**col2_cell.font.__dict__), Font  (**col1_cell.font.__dict__)
            col1_cell.fill, col2_cell.fill = PatternFill(**col2_cell.fill.__dict__),    PatternFill(**col1_cell.fill.__dict__)
            col1_cell.border, col2_cell.border = Border(**col2_cell.border.__dict__), Border    (**col1_cell.border.__dict__)

            # Apply grey shading based on yoy_quarterly_sales_growth value
            yoy_value = col2_value
            fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")     # Default fill color
            if yoy_value is not None and isinstance(yoy_value, (int, float)):
                if 10 <= yoy_value < 15:
                    fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")  # 5% grey
                elif 15 <= yoy_value < 20:
                    fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")  # 15% grey
                elif 20 <= yoy_value < 25:
                    fill = PatternFill(start_color="B3B3B3", end_color="B3B3B3", fill_type="solid")  # 25% grey
                elif yoy_value >= 25:
                    fill = PatternFill(start_color="999999", end_color="999999", fill_type="solid")  # 35% grey

            col1_cell.fill = fill


        # Save the modified Excel file
        output_filename = 'quarterly_sales_highight_nonbank.xlsx'
        wb.save(output_filename)
        print("File saved successfully.")

        # Email configuration
        smtp_host = os.getenv("SMTP_HOST")
        smtp_port = os.getenv("SMTP_PORT")
        sender_email = os.getenv("SENDER_EMAIL")
        sender_app_password = os.getenv("SENDER_APP_PASSWORD")  # You need to generate an App Password in your Gmail account settings
        recipient_email = os.getenv("RECIPIENT_EMAIL")

        # Create the email
        msg = MIMEMultipart()
        msg['From'] = sender_email
        msg['To'] = recipient_email
        msg['Subject'] = 'Highlight yoy_quarterly_sales_growth column in Non-Bank report.'

        # Attach the Excel file to the email
        with open(output_filename, "rb") as excel_file:
            part = MIMEApplication(excel_file.read(), Name=output_filename)
            part['Content-Disposition'] = f'attachment; filename="{output_filename}"'
            msg.attach(part)

        # Connect to the Gmail SMTP server and send the email
        with smtplib.SMTP(smtp_host, smtp_port) as server:
            server.starttls()
            server.login(sender_email, sender_app_password)
            server.sendmail(sender_email, recipient_email, msg.as_string())
        
        print("func30 executed successfully...")
    except Exception as e:
        print(f"Error : {e}")





def add_rank_columns(rank_instructions, stock_info_df):
    for col_name, order, rank_col_name in rank_instructions:
        if col_name in stock_info_df.columns:
            col_data = stock_info_df[col_name]
            mask = ~col_data.isnull()

            if order == 'asc':
                ranked_values = col_data[mask].rank(ascending=True, method='min')
            elif order == 'desc':
                ranked_values = col_data[mask].rank(ascending=False, method='min')

            rank_series = pd.Series(np.nan, index=stock_info_df.index)
            rank_series[mask] = ranked_values

            stock_info_df[rank_col_name] = rank_series

    # Calculate the overall_rank column as the mean of non-null rank columns
    non_null_rank_columns = [rank_col_name for col_name, _, rank_col_name in rank_instructions
                             if rank_col_name in stock_info_df.columns]
    stock_info_df['overall_rank'] = stock_info_df[non_null_rank_columns].mean(axis=1, skipna=True)

    return stock_info_df

# ['price_to_earning', 'asc', 'price_to_earning_rk'],

rank_instructions = [

    ['return_on_equity', 'desc', 'return_on_equity_rk'],
    ['price_to_earning', 'asc', 'price_to_earning_rk'],
    ['operating_profit_growth', 'desc', 'operating_profit_growth_rk'],
    ['yoy_quarterly_sales_growth', 'desc', 'yoy_quarterly_sales_growth_rk'],
    ['sales_growth', 'desc', 'sales_growth_rk'],
    ['sales_growth_3years', 'desc', 'sales_growth_3years_rk'],
    ['sales_growth_5years', 'desc', 'sales_growth_5years_rk'],
    ['sales_growth_7years', 'desc', 'sales_growth_7years_rk'],
    ['sales_growth_10years', 'desc', 'sales_growth_10years_rk'],
    # ['yoy_quarterly_profit_growth', 'desc', 'yoy_quarterly_profit_growth_rk'],
    # ['profit_growth', 'desc', 'profit_growth_rk'],
    # ['profit_growth_3years', 'desc', 'profit_growth_3years_rk'],
    # ['profit_growth_5years', 'desc', 'profit_growth_5years_rk'],
    # ['profit_growth_7years', 'desc', 'profit_growth_7years_rk'],
    # ['profit_growth_10years', 'desc', 'profit_growth_10years_rk'],
    # ['return_on_capital_employed', 'desc', 'return_on_capital_employed_rk'],
    # ['average_return_on_capital_employed_3years', 'desc', 'average_return_on_capital_employed_3years_rk'],
    # ['average_return_on_capital_employed_5years', 'desc', 'average_return_on_capital_employed_5years_rk'],
    # ['average_return_on_capital_employed_7years', 'desc', 'average_return_on_capital_employed_7years_rk'],
    # ['average_return_on_capital_employed_10years', 'desc', 'average_return_on_capital_employed_10years_rk'],
    ['opm', 'desc', 'opm_rk'],

]


rank_instructions_banks = [

    ['return_on_equity', 'desc', 'return_on_equity_rk'],
    ['return_on_assets', 'desc', 'return_on_assets_rk'],
    ['price_to_earning', 'asc', 'price_to_earning_rk'],
    ['price_to_book_value', 'asc', 'price_to_book_value_rk'],
    ['yoy_quarterly_sales_growth', 'desc', 'yoy_quarterly_sales_growth_rk'],
    ['sales_growth', 'desc', 'sales_growth_rk'],
    ['sales_growth_3years', 'desc', 'sales_growth_3years_rk'],
    ['sales_growth_5years', 'desc', 'sales_growth_5years_rk'],
    ['sales_growth_7years', 'desc', 'sales_growth_7years_rk'],
    ['sales_growth_10years', 'desc', 'sales_growth_10years_rk'],
    ['yoy_quarterly_profit_growth', 'desc', 'yoy_quarterly_profit_growth_rk'],
    ['profit_growth', 'desc', 'profit_growth_rk'],
    ['profit_growth_3years', 'desc', 'profit_growth_3years_rk'],
    ['profit_growth_5years', 'desc', 'profit_growth_5years_rk'],
    ['profit_growth_7years', 'desc', 'profit_growth_7years_rk'],
    ['profit_growth_10years', 'desc', 'profit_growth_10years_rk'],


]




excel_columns = [

    'name',
    'industry',
    'market_capitalization',
    'renumbered_rank',
    'overall_rank',
    'current_price',
    'return_over_1day',
    'return_over_1week',
    'return_over_1month',
    'return_over_3months',
    'return_over_1year',
    'return_over_3years',
    'return_over_5years',
    'return_over_7years',
    'return_over_10years',
    'return_on_equity',
    'return_on_assets',
    'operating_profit_growth',
    'yoy_quarterly_profit_growth',
    'profit_growth',
    'profit_growth_3years',
    'profit_growth_5years',
    'profit_growth_7years',
    'profit_growth_10years',
    'promoter_holding',
    'change_in_promoter_holding',
    'fii_holding',
    'change_in_fii_holding',
    'dii_holding',
    'dividend_yield',
    'yoy_quarterly_sales_growth',
    'sales_growth',
    'sales_growth_3years',
    'sales_growth_5years',
    'sales_growth_7years',
    'sales_growth_10years',
    'evebitda',
    'market_cap_to_sales',
    'price_to_earning',
    'return_on_capital_employed',
    'average_return_on_capital_employed_3years',
    'average_return_on_capital_employed_5years',
    'average_return_on_capital_employed_7years',
    'average_return_on_capital_employed_10years',
    'sales_growth_rk',
    'sales_growth_3years_rk',
    'sales_growth_5years_rk',
    'sales_growth_7years_rk',
    'sales_growth_10years_rk',
    'yoy_quarterly_sales_growth_rk',
    # 'profit_growth_rk',
    # 'profit_growth_3years_rk',
    # 'profit_growth_5years_rk',
    # 'profit_growth_7years_rk',
    # 'profit_growth_10years_rk',
    # 'return_on_capital_employed_rk',
    # 'average_return_on_capital_employed_3years_rk',
    # 'average_return_on_capital_employed_5years_rk',
    # 'average_return_on_capital_employed_7years_rk',
    # 'average_return_on_capital_employed_10years_rk',
    'return_on_equity_rk',
    'price_to_earning_rk',
    'operating_profit_growth_rk',
    'yoy_quarterly_sales_growth_rk',
    'opm_rk'

]


excel_columns_banks = [

    'name',
    'industry',
    'market_capitalization',
    'renumbered_rank',
    'overall_rank',
    'current_price',
    'return_over_1day',
    'return_over_1week',
    'return_over_1month',
    'return_over_3months',
    'return_over_1year',
    'return_over_3years',
    'return_over_5years',
    'return_over_7years',
    'return_over_10years',
    'return_on_equity',
    'return_on_assets',
    'price_to_book_value',
    'yoy_quarterly_profit_growth',
    'profit_growth',
    'profit_growth_3years',
    'profit_growth_5years',
    'profit_growth_7years',
    'profit_growth_10years',
    'promoter_holding',
    'change_in_promoter_holding',
    'fii_holding',
    'change_in_fii_holding',
    'dii_holding',
    'dividend_yield',
    'yoy_quarterly_sales_growth',
    'sales_growth',
    'sales_growth_3years',
    'sales_growth_5years',
    'sales_growth_7years',
    'sales_growth_10years',
    'price_to_earning',
    'sales_growth_rk',
    'sales_growth_3years_rk',
    'sales_growth_5years_rk',
    'sales_growth_7years_rk',
    'sales_growth_10years_rk',
    'yoy_quarterly_sales_growth_rk',
    'profit_growth_rk',
    'profit_growth_3years_rk',
    'profit_growth_5years_rk',
    'profit_growth_7years_rk',
    'profit_growth_10years_rk',
    'return_on_equity_rk',
    'price_to_earning_rk',
    'price_to_book_value_rk',
    'yoy_quarterly_sales_growth_rk',
    'return_on_assets_rk'

]



def func6(uploaddate,destination_table,cnx):

    try:
        engine = create_engine(f'mysql+mysqlconnector://{db_user}:{db_password}@{ip_addr}/{db_database}?connect_timeout=60')


        # Read data from the source table into a DataFrame
        query = f"SELECT * FROM {destination_table} WHERE uploaddate = '{uploaddate}'"
        df = pd.read_sql(query, con=engine)

        # selected_rows = df[
        #     ((df['sales_growth'] > 15) | (df['sales_growth_3years'] > 15) | (
        #                 df['sales_growth_5years'] > 15) | (df['sales_growth_7years'] > 15) | (
        #                  df['sales_growth_10years'] > 15))
        #     & (~df['industry'].str.lower().str.contains('finance|bank'))
        #     ]
        
        df['industry'] = df['industry'].fillna('Trading')
        selected_rows = df[
            ~df['industry'].str.lower().str.contains('finance|bank')
            ]


        df_ranked = add_rank_columns(rank_instructions, selected_rows)

        # Renumber the ranks based on the overall rank
        df_ranked['renumbered_rank'] = df_ranked['overall_rank'].rank(method='first')

        # Sort the DataFrame based on the renumbered ranks
        sorted_df = df_ranked.sort_values(by='renumbered_rank')
        

        # Construct the output Excel file name with the upload date
        output_filename = f'output_{uploaddate}.xlsx'
        sorted_df[excel_columns].to_excel(output_filename, index=False)

        # Email configuration
        smtp_host = os.getenv("SMTP_HOST")
        smtp_port = os.getenv("SMTP_PORT")
        sender_email = os.getenv("SENDER_EMAIL")
        sender_app_password = os.getenv("SENDER_APP_PASSWORD")  # You need to generate an App Password in your Gmail account settings
        recipient_email = os.getenv("RECIPIENT_EMAIL")
        
        

        # Create the email
        msg = MIMEMultipart()
        msg['From'] = sender_email
        msg['To'] = recipient_email
        msg['Subject'] = 'Non-Bank Data'

        # Attach the Excel file to the email
        with open(output_filename, "rb") as excel_file:
            part = MIMEApplication(excel_file.read(), Name=output_filename)
            part['Content-Disposition'] = f'attachment; filename="{output_filename}"'
            msg.attach(part)

        # Connect to the Gmail SMTP server and send the email
        with smtplib.SMTP(smtp_host, smtp_port) as server:
            server.starttls()
            server.login(sender_email, sender_app_password)
            server.sendmail(sender_email, recipient_email, msg.as_string())

    except Exception as e:
        print(f"Error: {e}")



def func7(uploaddate,destination_table,cnx):


    try:
        engine = create_engine(f'mysql+mysqlconnector://{db_user}:{db_password}@{ip_addr}/{db_database}?connect_timeout=60')


        # Read data from the source table into a DataFrame
        query = f"SELECT * FROM {destination_table} WHERE uploaddate = '{uploaddate}'"
        df = pd.read_sql(query, con=engine)

        # selected_rows = df[
        #     ((df['sales_growth'] > 15) | (df['sales_growth_3years'] > 15) | (
        #                 df['sales_growth_5years'] > 15) | (df['sales_growth_7years'] > 15) | (
        #                  df['sales_growth_10years'] > 15))
        #     & (df['industry'].str.lower().str.contains('finance|bank'))
        #     ]

        selected_rows = df[
            ((df['sales_growth'] > 15) | (df['sales_growth_3years'] > 15) | (
                    df['sales_growth_5years'] > 15) | (df['sales_growth_7years'] > 15) | (
                     df['sales_growth_10years'] > 15))
            & (df['industry'].str.lower().str.contains('finance|bank'))
            ]
        df_ranked = add_rank_columns(rank_instructions_banks, selected_rows)

        # Renumber the ranks based on the overall rank
        df_ranked['renumbered_rank'] = df_ranked['overall_rank'].rank(method='first')

        # Sort the DataFrame based on the renumbered ranks
        sorted_df = df_ranked.sort_values(by='renumbered_rank')


        # Construct the output Excel file name with the upload date
        output_filename = f'output_banks_{uploaddate}.xlsx'
        sorted_df[excel_columns_banks].to_excel(output_filename, index=False)

        # Email configuration
        smtp_host = os.getenv("SMTP_HOST")
        smtp_port = os.getenv("SMTP_PORT")
        sender_email = os.getenv("SENDER_EMAIL")
        sender_app_password = os.getenv("SENDER_APP_PASSWORD")  # You need to generate an App Password in your Gmail account settings
        recipient_email = os.getenv("RECIPIENT_EMAIL")
        

        # Create the email
        msg = MIMEMultipart()
        msg['From'] = sender_email
        msg['To'] = recipient_email
        msg['Subject'] = 'Bank Data'

        # Attach the Excel file to the email
        with open(output_filename, "rb") as excel_file:
            part = MIMEApplication(excel_file.read(), Name=output_filename)
            part['Content-Disposition'] = f'attachment; filename="{output_filename}"'
            msg.attach(part)

        # Connect to the Gmail SMTP server and send the email
        with smtplib.SMTP(smtp_host, smtp_port) as server:
            server.starttls()
            server.login(sender_email, sender_app_password)
            server.sendmail(sender_email, recipient_email, msg.as_string())

    except Exception as e:
        print(f"Error: {e}")

def func10 (nonbank,uploaddate):
    try:
        data = pd.read_excel(nonbank)
        data = pd.DataFrame(data)
        print(data.head())
        filtered_data = data[((data['change_in_promoter_holding'] > 0) |
                      (data['change_in_fii_holding'] > 0)) &
                      (data['return_over_3months'] < 20)]
        
        output_filename = f'promoter_holdings_nonbank_{uploaddate}.xlsx'
        filtered_data[excel_columns].to_excel(output_filename, index = False)



        # Email configuration
        smtp_host = os.getenv("SMTP_HOST")
        smtp_port = os.getenv("SMTP_PORT")
        sender_email = os.getenv("SENDER_EMAIL")
        sender_app_password = os.getenv("SENDER_APP_PASSWORD")  # You need to generate an App Password in your Gmail account settings
        recipient_email = os.getenv("RECIPIENT_EMAIL")
        

        # Create the email
        msg = MIMEMultipart()
        msg['From'] = sender_email
        msg['To'] = recipient_email
        msg['Subject'] = 'Non-Bank Promoter Holdings Data'

        # Attach the Excel file to the email
        with open(output_filename, "rb") as excel_file:
            part = MIMEApplication(excel_file.read(), Name=output_filename)
            part['Content-Disposition'] = f'attachment; filename="{output_filename}"'
            msg.attach(part)

        # Connect to the Gmail SMTP server and send the email
        with smtplib.SMTP(smtp_host, smtp_port) as server:
            server.starttls()
            server.login(sender_email, sender_app_password)
            server.sendmail(sender_email, recipient_email, msg.as_string())
        

    except Exception as e:
        print(f"Error: {e}")


def func11 (bank,uploaddate):
    try:
        data = pd.read_excel(bank)
        data = pd.DataFrame(data)
        print(data.head())

        filtered_data=data[((data['change_in_promoter_holding'] > 0) |
                      (data['change_in_fii_holding'] > 0)) &
                      (data['return_over_3months'] < 20)]
        
        print(filtered_data.shape)

        output_filename = f'promoter_holdings_bank_{uploaddate}.xlsx'
        filtered_data[excel_columns_banks].to_excel(output_filename, index = False)



        # Email configuration
        smtp_host = os.getenv("SMTP_HOST")
        smtp_port = os.getenv("SMTP_PORT")
        sender_email = os.getenv("SENDER_EMAIL")
        sender_app_password = os.getenv("SENDER_APP_PASSWORD")  # You need to generate an App Password in your Gmail account settings
        recipient_email = os.getenv("RECIPIENT_EMAIL")
        

        # Create the email
        msg = MIMEMultipart()
        msg['From'] = sender_email
        msg['To'] = recipient_email
        msg['Subject'] = 'Bank Promoter Holdings Data'

        # Attach the Excel file to the email
        with open(output_filename, "rb") as excel_file:
            part = MIMEApplication(excel_file.read(), Name=output_filename)
            part['Content-Disposition'] = f'attachment; filename="{output_filename}"'
            msg.attach(part)

        # Connect to the Gmail SMTP server and send the email
        with smtplib.SMTP(smtp_host, smtp_port) as server:
            server.starttls()
            server.login(sender_email, sender_app_password)
            server.sendmail(sender_email, recipient_email, msg.as_string())
        
        

    except Exception as e:
        print(f"Error: {e}")


def func12(nonbank, pnonbank,uploaddate):
    try:
        data1 = pd.read_excel(nonbank)
        data2 = pd.read_excel(pnonbank)

        wb = load_workbook(nonbank)
        ws = wb.active

        for i,row in enumerate(ws.iter_rows(min_row = 1, max_row = ws.max_row, min_col = 1,max_col = ws.max_column, values_only = True)):
            for cell_value in row:
                if cell_value in data2['name'].values:
                    for cell_to_format in ws[i+1]:
                        cell_to_format.fill = PatternFill(start_color = "e3edf6", end_color = "e3edf6", fill_type = "solid")
                    break

        output_filename = f"Highlight_output_nonbank_{uploaddate}.xlsx"
        wb.save(output_filename)
        print("Output file is highlighted.")
    except Exception as e:
        print(f"Error: {e}")
        

    
def func13(bank, pbank,uploaddate):
    try:
        data1 = pd.read_excel(bank)
        data2 = pd.read_excel(pbank)

        wb = load_workbook(bank)
        ws = wb.active

        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1,  max_col=ws.max_column, values_only=True)):
            for cell_value in row:
                if cell_value in data2['name'].values:
                    for cell_to_format in ws[i + 1]:
                        cell_to_format.fill = PatternFill(start_color="e3edf6",     end_color="e3edf6", fill_type="solid")
                    break
    
        output_filename = f"Highlight_output_bank_{uploaddate}.xlsx"
        wb.save(output_filename)
        print("Output file is highlighted.")

    except Exception as e:
        print("Error: {e}")
        
    

def func14(file_path, uploaddate):
    try:
        database_url = f'mysql+mysqlconnector://{db_user}:{db_password}@{ip_addr}/{dbname}'
        # Create an SQLAlchemy Engine with the URL
        engine = create_engine(database_url)

        print("connected successfully")

        # Read Excel file
        df_excel = pd.read_excel(file_path)

        wb = load_workbook(file_path)
        ws = wb.active

        # Iterate over each row in the DataFrame
        for index, row in df_excel.iterrows():
            name = row['name']
            # Query MySQL table for matching entries
            query = f'SELECT * FROM recommendations WHERE name = "{name}"'
            df_mysql = pd.read_sql(query, engine)

            # If there are matching entries, sort by 'recodate' and create comment string
            if not df_mysql.empty:
                df_mysql = df_mysql.sort_values(by='recodate', ascending=False)
                name = df_mysql['name'].tolist()
                recodates = df_mysql['recodate'].tolist()
                weightage = df_mysql['weightage'].tolist()
                addupto = df_mysql['addupto'].tolist()
                cmp = df_mysql['cmp'].tolist()
                sl = df_mysql['sl'].tolist()
                timeframe = df_mysql['timeframe'].tolist()
                targets = "\n".join(
                df_mysql[['target1', 'target2', 'target3', 'target4', 'target5', 'target6', 'target7', 'target8', 'target9']].astype(str).apply(lambda x: '/ '.join(x.dropna()), axis=1))

                targets = targets.split("\n")
                comment_text = ""
                for i in range(len(name)):
                    existing_text = f"Name: {name[i]},\nrecodate: {recodates[i]},\ncmp : {cmp[i]},\nWeightage: {weightage[i]},\nTimeframe: {timeframe[i]},\nAddupto : {addupto[i]},\nStopLoss : {sl[i]},\nTargets: {targets[i]}"
                    comment_text += existing_text + "\n\n"

                cell = ws.cell(row=index + 2, column=1)  # Since 'name' column is the first column

                comment = Comment(comment_text, "Author")
                comment_width = max(len(line) for line in comment_text.split('\n')) * 6
                comment_height = len(comment_text.split('\n')) * 20
                comment.width = comment_width if comment_width > 300 else 300
                comment.height = comment_height if comment_height > 200 else 200

                # Add comment to the cell
                cell.comment = comment

        # Save the updated Excel file
        output_filename = f'Commented_output_bank_{uploaddate}.xlsx'
        wb.save(output_filename)

        # Email configuration
        smtp_host = os.getenv("SMTP_HOST")
        smtp_port = os.getenv("SMTP_PORT")
        sender_email = os.getenv("SENDER_EMAIL")
        sender_app_password = os.getenv("SENDER_APP_PASSWORD")  # You need to generate anApp Password in your Gmail account settings
        recipient_email = os.getenv("RECIPIENT_EMAIL")
        

        # Create the email
        msg = MIMEMultipart()
        msg['From'] = sender_email
        msg['To'] = recipient_email
        msg['Subject'] = 'Commented bank Data'

        # Attach the Excel file to the email
        with open(output_filename, "rb") as excel_file:
            part = MIMEApplication(excel_file.read(), Name=output_filename)
            part['Content-Disposition'] = f'attachment; filename="{output_filename}"'
            msg.attach(part)

        # Connect to the Gmail SMTP server and send the email
        with smtplib.SMTP(smtp_host, smtp_port) as server:
            server.starttls()
            server.login(sender_email, sender_app_password)
            server.sendmail(sender_email, recipient_email, msg.as_string())
    except Exception as e:
        print(f"Error: {e}")


def func15(file_path,uploaddate):
    try:
        database_url = f'mysql+mysqlconnector://{db_user}:{db_password}@{ip_addr}/{dbname}'
        # Create an SQLAlchemy Engine with the URL
        engine = create_engine(database_url)
    
        print("connected successfully")

        # Read Excel file
        df_excel = pd.read_excel(file_path)

        wb = load_workbook(file_path)
        ws = wb.active

        # Iterate over each row in the DataFrame
        for index, row in df_excel.iterrows():
            name = row['name']
            # Query MySQL table for matching entries
            query = f'SELECT * FROM recommendations WHERE name = "{name}"'
            df_mysql = pd.read_sql(query, engine)

            # If there are matching entries, sort by 'recodate' and create comment string
            if not df_mysql.empty:
                df_mysql = df_mysql.sort_values(by='recodate', ascending=False)
                name = df_mysql['name'].tolist()
                recodates = df_mysql['recodate'].tolist()
                weightage = df_mysql['weightage'].tolist()
                addupto = df_mysql['addupto'].tolist()
                cmp = df_mysql['cmp'].tolist()
                sl = df_mysql['sl'].tolist()
                timeframe = df_mysql['timeframe'].tolist()
                targets = "\n".join(
                df_mysql[['target1', 'target2', 'target3', 'target4', 'target5', 'target6', 'target7', 'target8', 'target9']].astype(str).apply(lambda x: '/ '.join(x.dropna()), axis=1))

                targets = targets.split("\n")
                comment_text = ""
                for i in range(len(name)):
                    existing_text = f"Name: {name[i]},\nrecodate: {recodates[i]},\ncmp : {cmp[i]},\nWeightage: {weightage[i]},\nTimeframe: {timeframe[i]},\nAddupto : {addupto[i]},\nStopLoss : {sl[i]},\nTargets: {targets[i]}"
                    comment_text += existing_text + "\n\n"

                cell = ws.cell(row=index + 2, column=1)  # Since 'name' column is the first column

                comment = Comment(comment_text, "Author")
                comment_width = max(len(line) for line in comment_text.split('\n')) * 6
                comment_height = len(comment_text.split('\n')) * 20
                comment.width = comment_width if comment_width > 300 else 300
                comment.height = comment_height if comment_height > 200 else 200

                # Add comment to the cell
                cell.comment = comment

        # Save the updated Excel file
        output_filename = f'Commented_output_nonbank_{uploaddate}.xlsx'
        wb.save(output_filename)

        # Email configuration
        smtp_host = os.getenv("SMTP_HOST")
        smtp_port = os.getenv("SMTP_PORT")
        sender_email = os.getenv("SENDER_EMAIL")
        sender_app_password = os.getenv("SENDER_APP_PASSWORD")  # You need to generate anApp    Password in your Gmail account settings
        recipient_email = os.getenv("RECIPIENT_EMAIL")
        

        # Create the email
        msg = MIMEMultipart()
        msg['From'] = sender_email
        msg['To'] = recipient_email
        msg['Subject'] = 'Commented nonbank Data'

        # Attach the Excel file to the email
        with open(output_filename, "rb") as excel_file:
            part = MIMEApplication(excel_file.read(), Name=output_filename)
            part['Content-Disposition'] = f'attachment; filename="{output_filename}"'
            msg.attach(part)

        # Connect to the Gmail SMTP server and send the email
        with smtplib.SMTP(smtp_host, smtp_port) as server:
            server.starttls()
            server.login(sender_email, sender_app_password)
            server.sendmail(sender_email, recipient_email, msg.as_string())

    except Exception as e:
        print(f"Error: {e}")


def func8(uploaddate,csv_file_path):
    try:
        #To load the excelfile in python
        data=pd.read_csv(csv_file_path)
        data=pd.DataFrame(data)

        #To replace the space between two words in the column name
        data.columns=data.columns.str.replace(" ","_")
        #print(data.head())

        #To find the preveious day value
        M_cap_1day=data['Market_Capitalization']/(1+(data['Return_over_1day']/100))

        #To add a column 'Marketcap1day' in the data.
        data['Mcap1day']=M_cap_1day

        #To find the previous week value
        M_cap_1week=data['Market_Capitalization']/(1+(data['Return_over_1week']/100))
        data['Mcap1week']=M_cap_1week

        #To find the previous month value
        M_cap_1month=data['Market_Capitalization']/(1+(data['Return_over_1month']/100))
        data['Mcap1month']=M_cap_1month

        #To find the value of previous 3 months value
        data['Mcap3months']=data['Market_Capitalization']/(1+(data['Return_over_3months']/100))

        #To find the value of previous 6 months value
        data['Mcap6months']=data['Market_Capitalization']/(1+(data['Return_over_6months']/100))

        #To find the value of previous 1 year value
        data["Mcap1yr"]=data['Market_Capitalization']/(1+(data['Return_over_1year']/100))


        #To select neccessary columns
        interested_col=['Industry','Market_Capitalization','Mcap1day','Mcap1week',
                        'Mcap1month','Mcap3months','Mcap6months','Mcap1yr','Market_Capitalization_3years_back',
                       'Market_Capitalization_5years_back','Market_Capitalization_7years_back','Market_Capitalization_10years_back']

        new_df=data[interested_col]

        #To rename the column name
        new_col_name={"Market_Capitalization":"Mcap","Market_Capitalization_3years_back":"Mcap3yrs",
                      "Market_Capitalization_5years_back":"Mcap5yrs","Market_Capitalization_7years_back":"Mcap7yrs",
                     "Market_Capitalization_10years_back":"Mcap10yrs"}
        new_df=new_df.rename(columns=new_col_name)

        #To group the columns by industries and then round off it and reset index as 0,1,2,3,...
        new_df=new_df.groupby('Industry').sum().round().reset_index()

        #To detect percent change in Market_Capitalization and M_cap_1day
        new_df['%1day']=round(((new_df['Mcap']-new_df['Mcap1day'])/new_df['Mcap'])*100,2)

        #To detect the percent change in market capitalization and market capitalization previous 1 week
        new_df['%1week']=round(((new_df['Mcap']-new_df['Mcap1week'])/new_df['Mcap'])*100,2)

        #To detect the percent change in market capitalization and market captalization previous 1 month
        new_df['%1month']=round(((new_df['Mcap']-new_df['Mcap1month'])/new_df['Mcap'])*100,2)

        #To detect the percent change in market capitalization and market captalization previous 3 month
        new_df['%3months']=round(((new_df['Mcap']-new_df['Mcap3months'])/new_df['Mcap'])*100,2)

        #To detect the percent change in market capitalization and market captalization previous 6 month
        new_df['%6months']=round(((new_df['Mcap']-new_df['Mcap6months'])/new_df['Mcap'])*100,2)

        #To detect the percent change in market capitalization and market captalization previous 1 year
        new_df['%1yr']=round(((new_df['Mcap']-new_df['Mcap1yr'])/new_df['Mcap'])*100,2)

        #To detect the percent change in market capitalization and market captalization previous 3 years
        new_df['%3yrs']=round((((new_df['Mcap']/new_df['Mcap3yrs'])**(1/3))-1)*100,2)

        #To detect the percent change in market capitalization and market captalization previous 5 years
        new_df['%5yrs']=round((((new_df['Mcap']/new_df['Mcap5yrs'])**(1/5))-1)*100,2)

        #To detect the percent change in market capitalization and market captalization previous 7 years
        new_df['%7yrs']=round((((new_df['Mcap']/new_df['Mcap7yrs'])**(1/7))-1)*100,2)

        #To detect the percent change in market capitalization and market captalization previous 10 years
        new_df['%10yrs']=round((((new_df['Mcap']/new_df['Mcap10yrs'])**(1/10))-1)*100,2)

        #To change the order of columns
        desired_col_order=['Industry','Mcap','Mcap1day','%1day','Mcap1week','%1week',
                          'Mcap1month','%1month','Mcap3months','%3months',
                          'Mcap6months','%6months','Mcap1yr','%1yr',
                          'Mcap3yrs','%3yrs','Mcap5yrs','%5yrs',
                          'Mcap7yrs','%7yrs','Mcap10yrs','%10yrs']
        new_df=new_df[desired_col_order]

        #To store output in excel file
        output_filename=f'Market_cap_{uploaddate}.xlsx'
        new_df.to_excel(output_filename,index=False)

        # Email configuration
        smtp_host = os.getenv("SMTP_HOST")
        smtp_port = os.getenv("SMTP_PORT")
        sender_email = os.getenv("SENDER_EMAIL")
        sender_app_password = os.getenv("SENDER_APP_PASSWORD")  # You need to generate an App Password in your Gmail account settings
        recipient_email = os.getenv("RECIPIENT_EMAIL")
        
        

        # Create the email
        msg = MIMEMultipart()
        msg['From'] = sender_email
        msg['To'] = recipient_email
        msg['Subject'] = 'Market Capitalization Report'

        # Attach the Excel file to the email
        with open(output_filename, "rb") as excel_file:
            part = MIMEApplication(excel_file.read(), Name=output_filename)
            part['Content-Disposition'] = f'attachment; filename="{output_filename}"'
            msg.attach(part)

        # Connect to the Gmail SMTP server and send the email
        with smtplib.SMTP(smtp_host, smtp_port) as server:
            server.starttls()
            server.login(sender_email, sender_app_password)
            server.sendmail(sender_email, recipient_email, msg.as_string())
            print("mail sent successfully")


    except Exception as e:
        print(f"Error: {e}")

holding_columns=[
    #["name","name"],
    #["insudtry","industry"],
    ["market_capitalization","mcap"],
    ["yoy_quarterly_sales_growth","yoy_qtr_sg"],
    ["sales_growth","sg"],
    ["expected_quarterly_sales_growth","exp_qtr_sg"],
    ["sales_growth_3years","sg3"],
    ["sales_growth_5years","sg5"],
    ["sales_growth_10years_median","sg10median"],
    ["sales_growth_5years_median","sg5median"],
    ["sales_growth_7years","sg7"],
    ["sales_growth_10years","sg10"],
    ["opm_last_year","opm_last_yr"],
    ["opm_latest_quarter","opm_latest_qtr"],
    ["opm_preceding_year","opm_prec_yr"],
    ["opm_preceding_quarter","opm_prec_qtr"],
    ["opm_preceding_year_quarter","opm_prec_yr_qtr"],
    ["opm_5year","opm5"],
    ["opm_10year","opm10"],
    #["zerodha","zerodha"],
    ]


        
def func9(from_path1,to_path1,csv_file_path,distinct_instr_path,file1,file2,list14_path,holding_columns,holding_table,cnx):
    try:
        # Delete all data in the holding_table
        cursor = cnx.cursor()
        cursor.execute(f"DELETE FROM {holding_table}")
        cnx.commit()
        cursor.close()
        print(f"All data deleted from table '{holding_table}'")
       
        for file in os.listdir(from_path1):
            print(file)
            shutil.copy2(os.path.join(from_path1, file), os.path.join(to_path1, file))
            holding_account(os.path.join(to_path1, file), holding_table, cnx)
        print("All files copied")

        alldata=pd.DataFrame()
        #print(type(alldata))

        for file in os.listdir(to_path1):
            print(file)
            path=os.path.join(to_path1,file)
            file=pd.read_csv(path)
            temp_file=pd.DataFrame(file)
            temp_file['account_id']=os.path.basename(path)
            alldata=pd.concat([alldata,temp_file],ignore_index=True)
       
        print(alldata.head())
        
        alldata.columns=alldata.columns.str.lower().str.replace(" ","_")
        alldata.columns=alldata.columns.str.replace(".","_")
        #print(alldata.shape)

        #alldata.to_csv("ad.csv",index=False)

        data=pd.read_csv(csv_file_path)
        data=pd.DataFrame(data)
        data['zerodha']=None
        #print(data.head())
       
        #data=data.fillna('')
        #print(data.head())
       
        data.columns=data.columns.str.lower().str.replace(" ","_")
        data.columns=data.columns.str.replace(".","_")
        # Remove columns ending with _1, _2, _3, _4, _5, _6, _7, _8
        data = data.loc[:, ~data.columns.str.endswith(('_1', '_2', '_3', '_4', '_5', '_6', '_7', '_8'))]
        #print(data.shape)
       

        unique_instr=pd.read_csv(distinct_instr_path)
       
       
        unique_instr.columns = unique_instr.columns.str.lower().str.replace(' ', '_')
        unique_instr.columns = unique_instr.columns.str.replace('.', '_')
        unique_instr = unique_instr.loc[:, ~unique_instr.columns.str.endswith(('_1', '_2', '_3', '_4', '_5', '_6', '_7', '_8'))]

        print(unique_instr.head())
        print(data.head())
       
        #Create manually 'zerodha' column in merged_data
        #Iterate through unique_instr and update data based on matching 'bse_code'

        for i in unique_instr['bse_code']:
            if pd.notna(i):  # Check if bse_code is not null
                mask = data['bse_code'] == i
                if any(mask):
                    zerodha_value = unique_instr.loc[unique_instr['bse_code'] == i, 'zerodha_code'].values[0]
                    data.loc[mask, 'zerodha'] = zerodha_value

        # Display the modified data
        print(data.head())


        data1=pd.read_csv(list14_path)
        data1.columns=data1.columns.str.lower().str.replace(" ","_")
        #data1=data1.fillna('')
        #print(data1)

       

        # Iterate through data1 and update data based on matching 'nse_code' and 'zerodha_code'
        for zerodha_code_entry in data1['zerodha_code']:
            part_before_dash = zerodha_code_entry.split('-')[0]
            mask = data['nse_code'] == part_before_dash
            if any(mask):
                zerodha_value = data1.loc[data1['zerodha_code'] == zerodha_code_entry, 'zerodha_code'].values[0]
                data.loc[mask, 'zerodha'] = zerodha_value
        print(data.head())
        print("Null values are:",data['nse_code'].isnull().sum())
       
        mask = (data['nse_code'].notnull()) & (data['zerodha'].isnull())
        data.loc[mask, 'zerodha'] = data.loc[mask, 'nse_code']
       
        print("Merged Data is updated.")

        ######

        fdata=pd.merge(alldata,data[['name','industry','market_capitalization','yoy_quarterly_sales_growth','sales_growth','expected_quarterly_sales_growth',
                             'sales_growth_3years','sales_growth_5years','sales_growth_10years_median',
                              'sales_growth_5years_median','sales_growth_7years','sales_growth_10years','opm','opm_last_year','opm_latest_quarter',
                              'opm_preceding_year','opm_preceding_quarter','opm_preceding_year_quarter','opm_5year',
                             'opm_10year','zerodha']],left_on='instrument',right_on='zerodha',how='left')

        fdata['instrument'] = fdata['instrument'].combine_first(fdata['zerodha'])

        print("*********fdata will be:",fdata)

        #To assign rank
       
        # Merge data1 with data2 and data3 based on 'name' column
        file1=pd.read_excel(file1)
        file2=pd.read_excel(file2)
        merged = pd.merge(fdata, file1[['name', 'renumbered_rank']], on='name', how='left')
        merged = pd.merge(merged, file2[['name', 'renumbered_rank']], on='name', how='left')

        # Create a new 'rank' column in data1 using the 'renumbered_rank' values
        # If 'renumbered_rank' is present in both data2 and data3, prioritize data2
        merged['rank'] = merged['renumbered_rank_x'].combine_first(merged['renumbered_rank_y'])

        # Drop the additional 'renumbered_rank' columns
        merged = merged.drop(columns=['renumbered_rank_x', 'renumbered_rank_y'])

        # To select the top 5 companies which are above the holding company using rank

        # Create empty lists to store the names and ranks of companies
        top_company_info = [[] for _ in range(5)]

        # Iterate through each entry in the 'name' column of data1
        for index, row in merged.iterrows():
            name_entry = row['name']
            industry_entry = row['industry']
            rank_entry = row['rank']

            # Filter data2 based on the 'industry' column
            mask =industry_entry==file1['industry']
            if any(mask):
                industry_data = file1[file1['industry'] == industry_entry]
            else:
                industry_data = file2[file2['industry'] == industry_entry]
               
            #print(industry_data)

            # Create a mask to filter for renumbered_rank smaller than rank_entry
            mask = industry_data['renumbered_rank'] < rank_entry

            # Filter based on the mask and retrieve the 5 smallest renumbered_rank
            data4 = industry_data[mask].nsmallest(5, 'renumbered_rank')

            # Store the names and ranks in the corresponding lists
            names_list = list(data4['name'])
            ranks_list = list(data4['renumbered_rank'])

            # Extend the lists to have a length of 5
            names_list.extend([None] * (5 - len(names_list)))
            ranks_list.extend([None] * (5 - len(ranks_list)))

            for i, (name, rank) in enumerate(zip(names_list, ranks_list)):
                top_company_info[i].append(f"{name} ({rank})" if name is not None else None)

        # Create a new DataFrame with the top company names and ranks
        top_companies_df = pd.DataFrame({f'Top_{i + 1}': info for i, info in enumerate(top_company_info)})
        #print(top_companies_df)

        # Merge the top_companies_df with data1 based on the 'industry' column
        df1=pd.concat([merged, top_companies_df], axis=1)
        new_col={"market_capitalization":"mcap","yoy_quarterly_sales_growth":"yoy_qtr_sg","sales_growth":"sg","expected_quarterly_sales_growth":"exp_qtr_sg",
                "sales_growth_3years":"sg3","sales_growth_5years":"sg5","sales_growth_10years_median":"sg10median",
                "sales_growth_5years_median":"sg5median","sales_growth_7years":"sg7","sales_growth_10years":"sg10","opm_last_year":"opm_last_yr","opm_latest_quarter":"opm_latest_qtr",
                 "opm_preceding_year":"opm_prec_yr","opm_preceding_quarter":"opm_prec_qtr",
                 "opm_preceding_year_quarter":"opm_prec_yr_qtr","opm_5year":"opm5","opm_10year":"opm10"}
        df2=df1.rename(columns=new_col)



       
        desired_col_order=['instrument','qty_','avg__cost','ltp','cur__val','p&l','net_chg_','day_chg_','name','industry',
                          'mcap','yoy_qtr_sg','sg','exp_qtr_sg','sg3','sg5','sg10median','sg5median','sg7','sg10','opm',
                           'opm_last_yr','opm_latest_qtr','opm_prec_yr','opm_prec_qtr','opm_prec_yr_qtr','opm5','opm10','zerodha',
                          'rank','Top_1','Top_2','Top_3','Top_4','Top_5']
        df2=df2[desired_col_order]
       
        print(df2.head())
        print(df2.shape)

        #To make Unique alldata
        unique_alldata = alldata.groupby('instrument').first().reset_index()


        # Merge df2 with alldata based on the 'instrument' column
        df3 = pd.merge(df2, unique_alldata[['instrument', 'account_id']], how='left', on='instrument')
        

        # Reorder columns to include 'account_id' in the desired order
        desired_col_order_with_account_id = ['instrument', 'qty_', 'avg__cost', 'ltp', 'cur__val', 'p&l', 'net_chg_', 'day_chg_', 'name', 'industry',
                                              'mcap', 'yoy_qtr_sg', 'sg', 'exp_qtr_sg', 'sg3', 'sg5', 'sg10median', 'sg5median', 'sg7', 'sg10', 'opm',
                                              'opm_last_yr', 'opm_latest_qtr', 'opm_prec_yr', 'opm_prec_qtr', 'opm_prec_yr_qtr', 'opm5', 'opm10', 'zerodha',
                                              'rank', 'Top_1', 'Top_2', 'Top_3', 'Top_4', 'Top_5', 'account_id']

        df3 = df3[desired_col_order_with_account_id]

        
       
        for col_name,short_col_name in holding_columns:
            if short_col_name in df3.columns:
                col_data=df3[short_col_name]
            if col_name not in df3.columns:
                for i in df3['zerodha']:
                    mask= i==df3['zerodha']
                    if any(mask):
                        col_data=data.loc[data['zerodha']==i,col_name].values[0]
                        df3.loc[mask,col_name]=col_data
       

        
        # Save the updated data to a new CSV file
        output_filename=f'ranked_data.xlsx'
        df3.to_excel(output_filename, index=False)
        print("Data is written in the csv file")

        # Email configuration
        smtp_host = os.getenv("SMTP_HOST")
        smtp_port = os.getenv("SMTP_PORT")
        sender_email = os.getenv("SENDER_EMAIL")
        sender_app_password = os.getenv("SENDER_APP_PASSWORD")  # You need to generate an App Password in your Gmail account settings
        recipient_email = os.getenv("RECIPIENT_EMAIL")
        

        # Create the email
        msg = MIMEMultipart()
        msg['From'] = sender_email
        msg['To'] = recipient_email
        msg['Subject'] = 'Rank Report'

        # Attach the Excel file to the email
        with open(output_filename, "rb") as excel_file:
            part = MIMEApplication(excel_file.read(), Name=output_filename)
            part['Content-Disposition'] = f'attachment; filename="{output_filename}"'
            msg.attach(part)

        # Connect to the Gmail SMTP server and send the email
        with smtplib.SMTP(smtp_host, smtp_port) as server:
            server.starttls()
            server.login(sender_email, sender_app_password)
            server.sendmail(sender_email, recipient_email, msg.as_string())
            print("mail sent successfully")

    except Exception as e:
        print(f"Error: {e}")



def holding_account(file,table_name,cnx):
    try:
        df=pd.read_csv(file)
        print("read csv")
    
        # Convert column names to lowercase and replace spaces with underscores
        df.columns = df.columns.str.lower().str.replace(' ', '_')
        df.columns = df.columns.str.replace('.', '_')
        df.columns=df.columns.str.replace("&","_and_")

        # Remove any duplicate columns
        df = df.loc[:, ~df.columns.duplicated()]
    
        # Adding a column
        df['account_id']=os.path.basename(file)
    
        # Get the maximum length of each column
        column_lengths = df.astype(str).apply(lambda x: x.str.len()).max().tolist()

        # Limit column names to 64 characters
        df.columns = df.columns.str[:64]

        # Convert numeric columns to float and replace empty or null values with NaN
        numeric_cols = df.select_dtypes(include=[np.number]).columns
        df[numeric_cols] = df[numeric_cols].apply(pd.to_numeric, errors='coerce')
    
    
        # Create the stockinfo table in MySQL
        cursor = cnx.cursor()
    

        create_table_query = f"CREATE TABLE IF NOT EXISTS {table_name} ("
        for col, length in zip(df.columns, column_lengths):
            col_data_type = 'FLOAT' if col in numeric_cols else f"VARCHAR({length})"
            create_table_query += f"{col} {col_data_type}, "
        create_table_query = create_table_query.rstrip(', ')
        create_table_query+=")"
        
        cursor.execute(create_table_query)
    
    
        # Insert data into the stockinfo table
        for row in df.itertuples(index=False, name=None):
            row_without_nan = [value if pd.notna(value) else None for value in row]
            insert_query = f"INSERT INTO {table_name} VALUES ({', '.join(['%s'] * len(row_without_nan))})"
            cursor.execute(insert_query, row_without_nan)
    
    
    
        # Commit changes and close the connection
        cnx.commit()
        cursor.close()

        print(f"Stockinfo table '{table_name}' created successfully.")

    except Exception as e:
        print(f"Error: {e}")

            
                    

# Specify the path to the CSV file and the table name
csv_file = os.getenv("CSV_FILE")
# print(csv_file)
dateObj = time.strftime("%Y%m%d")
#dateObj="20230907"

to_path = os.environ["TO_PATH"]
to_path = to_path + "\\" + dateObj

from_path=os.getenv("FROM_PATH")

table_name = os.getenv("TABLE_NAME")+ dateObj
destination_table = os.getenv("DESTINATION_TABLE")
merged_path=os.path.join(to_path, csv_file)

from_path1= os.getenv("FROM_PATH1")
to_path1=os.getenv("TO_PATH1")
distinct_instr_path=os.getenv("DISTINCT_INSTR_PATH")
list14_path=os.getenv("LIST14_PATH")

nonbank=f"{os.getenv('TO_PATH')}" + "\\" + f"output_{dateObj}.xlsx"

bank=f"{os.getenv('TO_PATH')}" + "\\" + f"output_banks_{dateObj}.xlsx"

promoter_bank = f"{os.getenv('TO_PATH')}" + "\\" +  f"promoter_holdings_bank_{dateObj}.xlsx"

promoter_nonbank = f"{os.getenv('TO_PATH')}" + "\\" +  f"promoter_holdings_nonbank_{dateObj}.xlsx"


Highlight_bank = f"{os.getenv('TO_PATH')}" + "\\" + f"Highlight_output_bank_{dateObj}.xlsx"
Highlight_nonbank = f"{os.getenv('TO_PATH')}" + "\\" + f"Highlight_output_nonbank_{dateObj}.xlsx"

commented_bank = f"{os.getenv('TO_PATH')}" + "\\" + f"commented_output_bank_{dateObj}.xlsx"
commented_nonbank = f"{os.getenv('TO_PATH')}" + "\\" + f"commented_output_nonbank_{dateObj}.xlsx"

holding_bank_file = f"{os.getenv('TO_PATH')}" + "\\" + f"holding_highlighted_comment_bank_{dateObj}.xlsx"
holding_nonbank_file = f"{os.getenv('TO_PATH')}" + "\\" + f"holding_highlighted_comment_nonbank_{dateObj}.xlsx"

volume_commented_bank_file = f"{os.getenv('TO_PATH')}" + "\\" + f"Volume_commented_bank_{dateObj}.xlsx"

volume_commented_nonbank_file = f"{os.getenv('TO_PATH')}" + "\\" + f"Volume_commented_nonbank_{dateObj}.xlsx"

volume_fno_bank = f"{os.getenv('TO_PATH')}" + "\\" + f"volume_fno_commented_bank.xlsx"
volume_fno_nonbank = f"{os.getenv('TO_PATH')}" + "\\" + f"volume_fno_commented_nonbank.xlsx"

summary_table = os.getenv("SUMMARY_TABLE")
v_info_table = os.getenv('V_INFO_TABLE')
holding_table = os.getenv('HOLDING_TABLE')
data_of_139 = os.getenv('DATA_OF_139')
nse_bse_table = os.getenv('NSE_BSE_TABLE')
fnotbale = os.getenv('FNO_TABLE')
fnolist = os.getenv('FNO_LIST')
#DownloadData
#func1(to_path)

#saveFilesToPath
#func2(from_path,to_path)

#mergeCSVFiles
#func3(to_path, csv_file)

# Create the stockinfo table in MySQL
#create_stockinfo_table
#func4(merged_path, table_name)


#copy_data_from_table
#func5(table_name, destination_table,dateObj)


#generateRankFile
#func6(dateObj,destination_table)

# func8: generate market capitalization report

#func9 : generate the report of top 5 companies

#func16(table_name,cnx)
#create new table account_summary in MySQL

#func17:  Create new table 'nse_code_and_bse_code' in mysql

#func18: update the nse_code column in stockinfo table from the 'nse_code_and_bse_code' table
# Create an index on the name column to speed up the update process

#update stockinfo table: Updating holding_summary column in stockinfo
#func19('account_summary','stockinfo',cnx,uploaddate)

#func20: Extract the list of stocks which are present in instrument column of account_summary but not in nse_code column of stockinfo

#func21 and func22: highlight and add comment on industry which are present in holdings

#func23 : Insert the data in v_info_table

#func24 and func25: Add and highlight the market_capitalization column

#func26 : create new table fnotable in MySQL.

#func27 and func28 : Add new fno column

# func29 and func30 : rearrange two columns and color the yoy_quarterly_sales_growth column.

def main():
    args = sys.argv[1:]

    if not args:
        func1(to_path)
        func2(from_path,to_path)
        func3(to_path, csv_file)
        func4(merged_path, table_name,cnx)
        func5(table_name, destination_table,dateObj,cnx,size=100)
        func6(dateObj,destination_table,cnx)
        func7(dateObj, destination_table,cnx)
        func8(dateObj,merged_path)
        func9(from_path1,to_path1,merged_path,distinct_instr_path,nonbank,bank,list14_path,holding_columns,holding_table,cnx2)
        func10(nonbank,dateObj)
        func11(bank,dateObj)
        func12(nonbank, promoter_nonbank,dateObj)
        func13(bank,promoter_bank,dateObj)
        func14(Highlight_bank,dateObj)
        func15(Highlight_nonbank,dateObj)
        func16(summary_table,cnx)
        func17(data_of_139,nse_bse_table,cnx)
        func18(destination_table,nse_bse_table,cnx,dateObj)
        func19(summary_table,destination_table,cnx,dateObj)
        func20(summary_table,destination_table,cnx,dateObj)
        func21(commented_bank,destination_table,dateObj)
        func22(commented_nonbank,destination_table,dateObj)
        func23(merged_path,v_info_table,dateObj,cnx1)
        func24(holding_bank_file,destination_table,cnx,dateObj)  
        func25(holding_nonbank_file, destination_table, cnx, dateObj) 
        func26(fnolist,fnotbale,cnx)
        func27(fnotbale, destination_table,volume_commented_bank_file, cnx,dateObj)
        func28(fnotbale, destination_table,volume_commented_nonbank_file, cnx,dateObj)
        func29(volume_fno_bank)
        func30(volume_fno_nonbank)
    else:
        try:
            selected_function = int(args[0])
            if selected_function == 1:
                func1(to_path)
                func2(from_path,to_path)
                func3(to_path, csv_file)
                func4(merged_path, table_name,cnx)
                func5(table_name, destination_table,dateObj,cnx,size=100)
                func6(dateObj,destination_table,cnx)
                func7(dateObj, destination_table,cnx)
                func8(dateObj,merged_path)
                func9(from_path1,to_path1,merged_path,distinct_instr_path,nonbank,bank,   list14_path,holding_columns,holding_table,cnx2)
                func10(nonbank,dateObj)
                func11(bank,dateObj)
                func12(nonbank, promoter_nonbank,dateObj)
                func13(bank,promoter_bank,dateObj)
                func14(Highlight_bank,dateObj)
                func15(Highlight_nonbank,dateObj)
                func16(summary_table,cnx)
                func17(data_of_139,nse_bse_table,cnx)
                func18(destination_table,nse_bse_table,cnx,dateObj)
                func19(summary_table,destination_table,cnx,dateObj)
                func20(summary_table,destination_table,cnx,dateObj)
                func21(commented_bank,destination_table,dateObj)
                func22(commented_nonbank,destination_table,dateObj)
                func23(merged_path,v_info_table,dateObj,cnx1)       
            elif selected_function == 2:
                func2(from_path,to_path)
                func3(to_path, csv_file)
                func4(merged_path, table_name,cnx)
                func5(table_name, destination_table,dateObj,cnx,size=100)
                func6(dateObj,destination_table,cnx)
                func7(dateObj, destination_table,cnx)
                func8(dateObj,merged_path)
                func9(from_path1,to_path1,merged_path,distinct_instr_path,nonbank,bank,   list14_path,holding_columns,holding_table,cnx2)
                func10(nonbank,dateObj)
                func11(bank,dateObj)
                func12(nonbank, promoter_nonbank,dateObj)
                func13(bank,promoter_bank,dateObj)
                func14(Highlight_bank,dateObj)
                func15(Highlight_nonbank,dateObj)
                func16(summary_table,cnx)
                func17(data_of_139,nse_bse_table,cnx)
                func18(destination_table,nse_bse_table,cnx,dateObj)
                func19(summary_table,destination_table,cnx,dateObj)
                func20(summary_table,destination_table,cnx,dateObj)
                func21(commented_bank,destination_table,dateObj)
                func22(commented_nonbank,destination_table,dateObj)
                func23(merged_path,v_info_table,dateObj,cnx1)       
            elif selected_function == 3:
                func3(to_path, csv_file)
                func4(merged_path, table_name,cnx)
                func5(table_name, destination_table,dateObj,cnx,size=100)
                func6(dateObj,destination_table,cnx)
                func7(dateObj, destination_table,cnx)
                func8(dateObj,merged_path)
                func9(from_path1,to_path1,merged_path,distinct_instr_path,nonbank,bank,   list14_path,holding_columns,holding_table,cnx2)
                func10(nonbank,dateObj)
                func11(bank,dateObj)
                func12(nonbank, promoter_nonbank,dateObj)
                func13(bank,promoter_bank,dateObj)
                func14(Highlight_bank,dateObj)
                func15(Highlight_nonbank,dateObj)
                func16(summary_table,cnx)
                func17(data_of_139,nse_bse_table,cnx)
                func18(destination_table,nse_bse_table,cnx,dateObj)
                func19(summary_table,destination_table,cnx,dateObj)
                func20(summary_table,destination_table,cnx,dateObj)
                func21(commented_bank,destination_table,dateObj)
                func22(commented_nonbank,destination_table,dateObj)
                func23(merged_path,v_info_table,dateObj,cnx1)       
            elif selected_function == 4:
                func4(merged_path, table_name,cnx)
                func5(table_name, destination_table,dateObj,cnx,size=100)
                func6(dateObj,destination_table,cnx)
                func7(dateObj, destination_table,cnx)
                func8(dateObj,merged_path)
                func9(from_path1,to_path1,merged_path,distinct_instr_path,nonbank,bank,   list14_path,holding_columns,holding_table,cnx2)
                func10(nonbank,dateObj)
                func11(bank,dateObj)
                func12(nonbank, promoter_nonbank,dateObj)
                func13(bank,promoter_bank,dateObj)
                func14(Highlight_bank,dateObj)
                func15(Highlight_nonbank,dateObj)
                func16(summary_table,cnx)
                func17(data_of_139,nse_bse_table,cnx)
                func18(destination_table,nse_bse_table,cnx,dateObj)
                func19(summary_table,destination_table,cnx,dateObj)
                func20(summary_table,destination_table,cnx,dateObj)
                func21(commented_bank,destination_table,dateObj)
                func22(commented_nonbank,destination_table,dateObj)
                func23(merged_path,v_info_table,dateObj,cnx1)       
            elif selected_function == 5:
                func5(table_name, destination_table,dateObj,cnx,size=100)
                func6(dateObj,destination_table,cnx)
                func7(dateObj, destination_table,cnx)
                func8(dateObj,merged_path)
                func9(from_path1,to_path1,merged_path,distinct_instr_path,nonbank,bank,   list14_path,holding_columns,holding_table,cnx2)
                func10(nonbank,dateObj)
                func11(bank,dateObj)
                func12(nonbank, promoter_nonbank,dateObj)
                func13(bank,promoter_bank,dateObj)
                func14(Highlight_bank,dateObj)
                func15(Highlight_nonbank,dateObj)
                func16(summary_table,cnx)
                func17(data_of_139,nse_bse_table,cnx)
                func18(destination_table,nse_bse_table,cnx,dateObj)
                func19(summary_table,destination_table,cnx,dateObj)
                func20(summary_table,destination_table,cnx,dateObj)
                func21(commented_bank,destination_table,dateObj)
                func22(commented_nonbank,destination_table,dateObj)
                func23(merged_path,v_info_table,dateObj,cnx1)       
            elif selected_function == 6:
                func6(dateObj,destination_table,cnx)
                func7(dateObj, destination_table,cnx)
                func8(dateObj,merged_path)
                func9(from_path1,to_path1,merged_path,distinct_instr_path,nonbank,bank,list14_path,holding_columns,holding_table,cnx2)
                func10(nonbank,dateObj)
                func11(bank,dateObj)
                func12(nonbank, promoter_nonbank,dateObj)
                func13(bank,promoter_bank,dateObj)
                func14(Highlight_bank,dateObj)
                func15(Highlight_nonbank,dateObj)
                func16(summary_table,cnx)
                func17(data_of_139,nse_bse_table,cnx)
                func18(destination_table,nse_bse_table,cnx,dateObj)
                func19(summary_table,destination_table,cnx,dateObj)
                func20(summary_table,destination_table,cnx,dateObj)
                func21(commented_bank,destination_table,dateObj)
                func22(commented_nonbank,destination_table,dateObj)
                func23(merged_path,v_info_table,dateObj,cnx1)
                func24(holding_bank_file,destination_table,cnx,dateObj)  
                func25(holding_nonbank_file, destination_table, cnx, dateObj) 
                func26(fnolist,fnotbale,cnx)
                func27(fnotbale, destination_table,volume_commented_bank_file, cnx,dateObj)
                func28(fnotbale, destination_table,volume_commented_nonbank_file, cnx,dateObj)
                func29(volume_fno_bank)
                func30(volume_fno_nonbank)

            elif selected_function == 8:
                func8(dateObj,merged_path)
                func9(from_path1,to_path1,merged_path,distinct_instr_path,nonbank,bank,list14_path,holding_columns,holding_table,cnx2)
                func10(nonbank,dateObj)
                func11(bank,dateObj)
                func12(nonbank, promoter_nonbank,dateObj)
                func13(bank,promoter_bank,dateObj)
                func14(Highlight_bank,dateObj)
                func15(Highlight_nonbank,dateObj)
            elif selected_function == 9:
                func9(from_path1,to_path1,merged_path,distinct_instr_path,nonbank,bank,list14_path,holding_columns,holding_table,cnx2)
                func10(nonbank,dateObj)
                func11(bank,dateObj)
                func12(nonbank, promoter_nonbank,dateObj)
                func13(bank,promoter_bank,dateObj)
                func14(Highlight_bank,dateObj)
                func15(Highlight_nonbank,dateObj)
                func16(summary_table,cnx)
                func19(summary_table,destination_table,cnx,dateObj)
                func21(commented_bank,destination_table,dateObj)
                func22(commented_nonbank,destination_table,dateObj)
                func23(merged_path,v_info_table,dateObj,cnx1)
            elif selected_function == 10:
                func10(nonbank,dateObj)
                func11(bank,dateObj)
                func12(nonbank, promoter_nonbank,dateObj)
                func13(bank,promoter_bank,dateObj)
                func14(Highlight_bank,dateObj)
                func15(Highlight_nonbank,dateObj)
            elif selected_function == 12:
                func12(nonbank, promoter_nonbank,dateObj)
                func13(bank,promoter_bank,dateObj)
                func14(Highlight_bank,dateObj)
                func15(Highlight_nonbank,dateObj)
            elif selected_function == 13:
                func13(bank,promoter_bank,dateObj)
                func14(Highlight_bank,dateObj)
                func15(Highlight_nonbank,dateObj)
            elif selected_function == 14:
                func14(Highlight_bank,dateObj)
                func15(Highlight_nonbank,dateObj)
            elif selected_function == 16:
                func16(summary_table,cnx)
                func17(data_of_139,nse_bse_table,cnx)
                func18(destination_table,nse_bse_table,cnx,dateObj)
                func19(summary_table,destination_table,cnx,dateObj)
                func20(summary_table,destination_table,cnx,dateObj)
                func21(commented_bank,destination_table,dateObj)
                func22(commented_nonbank,destination_table,dateObj)
                func23(merged_path,v_info_table,dateObj,cnx1)       
            elif selected_function == 17:
                func17(data_of_139,nse_bse_table,cnx)
                func18(destination_table,nse_bse_table,cnx,dateObj)
                func19(summary_table,destination_table,cnx,dateObj)
                func20(summary_table,destination_table,cnx,dateObj)
                func21(commented_bank,destination_table,dateObj)
                func22(commented_nonbank,destination_table,dateObj)
                func23(merged_path,v_info_table,dateObj,cnx1)    
            elif selected_function == 24:
                func24(holding_bank_file,destination_table,cnx,dateObj)  
                func25(holding_nonbank_file, destination_table, cnx, dateObj) 
            elif selected_function == 26:
                # func26(fnolist,fnotbale,cnx)
                func27(fnotbale, destination_table,volume_commented_bank_file, cnx,dateObj)
                func28(fnotbale, destination_table,volume_commented_nonbank_file, cnx,dateObj)
                func29(volume_fno_bank)
                func30(volume_fno_nonbank)
            elif selected_function == 29:
                # func29(volume_fno_bank)
                # func30(volume_fno_nonbank)
                func29(volume_fno_bank,cnx)
            else:
                print("Invalid input. Please provide a number between 1 and 17.")
        except ValueError:
            print("Invalid input. Please provide a number between 1 and 17.")


# Don't forget to close the connection when you're done
            cnx.close()


if __name__ == "__main__":
    main()
